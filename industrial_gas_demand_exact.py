#!/usr/bin/env python3
"""
Exact replication of the "Industrial gas demand" sheet from Excel.
Uses precise 3-criteria SUMIFS logic matching the Excel formulas.
"""

import pandas as pd
import numpy as np
import openpyxl
from datetime import datetime
import logging
import warnings
warnings.filterwarnings('ignore')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def load_multiticker_with_full_metadata(file_path='use4.xlsx', sheet_name='MultiTicker'):
    """
    Load MultiTicker data with complete metadata from rows 14, 15, 16.
    """
    logger.info(f"Loading MultiTicker with full metadata from {file_path}")
    
    # Load workbook to extract metadata
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    
    # Extract metadata from rows 14 (category), 15 (region), 16 (subcategory)
    metadata = {}
    max_col = min(ws.max_column, 600)  # Extended range
    
    logger.info(f"Extracting metadata from columns C to {openpyxl.utils.get_column_letter(max_col)}")
    
    for col in range(3, max_col + 1):  # Starting from column C
        col_name = f'Col_{col-2}'  # Col_1, Col_2, etc.
        category = ws.cell(row=14, column=col).value
        region = ws.cell(row=15, column=col).value  
        subcategory = ws.cell(row=16, column=col).value
        
        metadata[col_name] = {
            'category': str(category) if category else '',
            'region': str(region) if region else '',
            'subcategory': str(subcategory) if subcategory else ''
        }
    
    wb.close()
    
    # Load data using pandas
    df_full = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    
    # Data starts from row 21 (index 20), column B onwards
    data_rows = df_full.iloc[20:, 1:max_col].copy()
    
    # Set column names: Date + Col_1, Col_2, etc.
    data_rows.columns = ['Date'] + [f'Col_{i}' for i in range(1, len(data_rows.columns))]
    
    # Convert Date column
    data_rows['Date'] = pd.to_datetime(data_rows['Date'], errors='coerce')
    
    # Remove invalid dates
    data_rows = data_rows.dropna(subset=['Date'])
    
    # Convert data columns to numeric
    for col in data_rows.columns[1:]:
        data_rows[col] = pd.to_numeric(data_rows[col], errors='coerce')
    
    logger.info(f"Loaded {len(data_rows)} dates with {len(metadata)} tickers")
    
    return data_rows, metadata


def sumifs_three_criteria(data_df, metadata, category_target, region_target, subcategory_target):
    """
    Exact replication of Excel 3-criteria SUMIFS:
    =SUMIFS(MultiTicker!$C21:$PB21,
            MultiTicker!$C$14:$PB$14, category_target,
            MultiTicker!$C$15:$PB$15, region_target,
            MultiTicker!$C$16:$PB$16, subcategory_target)
    """
    matching_cols = []
    
    for col, info in metadata.items():
        # Exact string matching (case sensitive as in Excel)
        category_match = info['category'] == category_target
        region_match = info['region'] == region_target
        subcategory_match = info['subcategory'] == subcategory_target
        
        if category_match and region_match and subcategory_match:
            matching_cols.append(col)
    
    if not matching_cols:
        logger.debug(f"No matches for {category_target}/{region_target}/{subcategory_target}")
        return pd.Series(0.0, index=data_df.index)
    
    logger.debug(f"Found {len(matching_cols)} matches for {category_target}/{region_target}/{subcategory_target}")
    
    # Sum across matching columns (handling NaN as 0)
    result = data_df[matching_cols].sum(axis=1, skipna=True)
    
    return result


def create_industrial_demand_sheet_exact(data_df, metadata):
    """
    Create exact replication of Industrial gas demand sheet structure.
    
    Columns:
    C: France Industrial
    D: Belgium Industrial  
    E: Italy Industrial
    F: GB Industrial
    G: Netherlands Industrial and Power
    H: Netherlands Zebra
    I: Germany Industrial and Power (total)
    J: Germany Gas-to-Power (to subtract)
    K: Germany Industrial (calculated = I - J)
    L: Total Industrial (sum of all components)
    """
    logger.info("Creating exact Industrial gas demand sheet")
    
    result = pd.DataFrame()
    result['Date'] = data_df['Date']
    
    # Column C: France Industrial
    # Criteria: Demand / France / Industrial
    result['France_Industrial'] = sumifs_three_criteria(
        data_df, metadata, 'Demand', 'France', 'Industrial'
    )
    
    # Column D: Belgium Industrial  
    # Criteria: Demand / Belgium / Industrial
    result['Belgium_Industrial'] = sumifs_three_criteria(
        data_df, metadata, 'Demand', 'Belgium', 'Industrial'
    )
    
    # Column E: Italy Industrial
    # Criteria: Demand / Italy / Industrial  
    result['Italy_Industrial'] = sumifs_three_criteria(
        data_df, metadata, 'Demand', 'Italy', 'Industrial'
    )
    
    # Column F: GB Industrial
    # Criteria: Demand / GB / Industrial
    result['GB_Industrial'] = sumifs_three_criteria(
        data_df, metadata, 'Demand', 'GB', 'Industrial'
    )
    
    # Column G: Netherlands Industrial and Power
    # Criteria: Demand / Netherlands / Industrial and Power
    result['Netherlands_IndPower'] = sumifs_three_criteria(
        data_df, metadata, 'Demand', 'Netherlands', 'Industrial and Power'
    )
    
    # Column H: Netherlands Zebra
    # Criteria: Demand / Netherlands / Zebra
    result['Netherlands_Zebra'] = sumifs_three_criteria(
        data_df, metadata, 'Demand', 'Netherlands', 'Zebra'
    )
    
    # Column I: Germany Industrial and Power (total)
    # Criteria: Demand / Germany / Industrial and Power
    result['Germany_Total'] = sumifs_three_criteria(
        data_df, metadata, 'Demand', 'Germany', 'Industrial and Power'
    )
    
    # Column J: Germany Gas-to-Power (to subtract)
    # Criteria: Intermediate Calculation / #Germany / Gas-to-Power
    result['Germany_GtP'] = sumifs_three_criteria(
        data_df, metadata, 'Intermediate Calculation', '#Germany', 'Gas-to-Power'
    )
    
    # Column K: Germany Industrial (calculated = I - J)
    result['Germany_Industrial'] = result['Germany_Total'] - result['Germany_GtP']
    
    # Column L: Total Industrial = SUM(C:H) + K
    # Sum of all industrial components
    industrial_components = [
        'France_Industrial', 'Belgium_Industrial', 'Italy_Industrial', 'GB_Industrial',
        'Netherlands_IndPower', 'Netherlands_Zebra', 'Germany_Industrial'
    ]
    
    result['Total_Industrial_Demand'] = result[industrial_components].sum(axis=1)
    
    # Sort by date
    result = result.sort_values('Date').reset_index(drop=True)
    
    logger.info(f"Created Industrial demand sheet with {len(result)} rows")
    
    return result


def validate_industrial_exact(industrial_df):
    """
    Validate against exact Excel values.
    """
    logger.info("Validating against Excel Industrial gas demand sheet")
    
    # Known exact values from Excel
    validations = [
        {
            'date': '2016-10-03',
            'France_Industrial': 31.903467,
            'Belgium_Industrial': 9.667990,
            'Italy_Industrial': 32.627039,
            'GB_Industrial': 10.210000,
            'Netherlands_IndPower': 42.541229,
            'Netherlands_Zebra': 3.373815,
            'Germany_Total': 114.791154,
            'Germany_GtP': 16.314490,
            'Germany_Industrial': 98.476664,
            'Total_Industrial_Demand': 228.800203
        }
    ]
    
    for val_data in validations:
        date = pd.to_datetime(val_data['date'])
        row = industrial_df[industrial_df['Date'] == date]
        
        if row.empty:
            logger.warning(f"Date {val_data['date']} not found")
            continue
        
        logger.info(f"\nValidation for {val_data['date']}:")
        
        all_match = True
        for col, expected in val_data.items():
            if col == 'date':
                continue
                
            actual = row[col].iloc[0]
            diff = abs(actual - expected)
            
            if diff < 0.01:
                logger.info(f"  ✓ {col}: {actual:.6f} (expected {expected:.6f})")
            else:
                logger.warning(f"  ✗ {col}: {actual:.6f} (expected {expected:.6f}, diff: {diff:.6f})")
                all_match = False
        
        if all_match:
            logger.info(f"  ✅ All values match for {val_data['date']}")
        else:
            logger.warning(f"  ⚠️  Some values don't match for {val_data['date']}")


def export_industrial_demand(industrial_df, output_file='industrial_demand_exact.csv'):
    """
    Export Industrial demand data matching Excel format.
    """
    # Format date column
    export_df = industrial_df.copy()
    export_df['Date'] = export_df['Date'].dt.strftime('%Y-%m-%d')
    
    # Round to match Excel precision
    numeric_cols = export_df.select_dtypes(include=[np.number]).columns
    export_df[numeric_cols] = export_df[numeric_cols].round(6)
    
    # Export full detail
    export_df.to_csv(output_file, index=False)
    logger.info(f"Exported detailed industrial data to {output_file}")
    
    # Also export just the total for Daily historic data integration
    total_only = export_df[['Date', 'Total_Industrial_Demand']].copy()
    total_only.to_csv('total_industrial_demand.csv', index=False)
    logger.info(f"Exported Total_Industrial_Demand to total_industrial_demand.csv")
    
    return output_file


def main():
    """
    Main execution function.
    """
    try:
        # Load MultiTicker with full metadata
        data_df, metadata = load_multiticker_with_full_metadata('use4.xlsx', 'MultiTicker')
        
        # Create exact Industrial demand sheet
        industrial_df = create_industrial_demand_sheet_exact(data_df, metadata)
        
        # Validate against Excel
        validate_industrial_exact(industrial_df)
        
        # Display key results
        logger.info(f"\nIndustrial demand for 2016-10-03:")
        sample = industrial_df[industrial_df['Date'] == '2016-10-03']
        if not sample.empty:
            logger.info(f"  Total Industrial: {sample['Total_Industrial_Demand'].iloc[0]:.6f}")
            logger.info(f"  France: {sample['France_Industrial'].iloc[0]:.6f}")
            logger.info(f"  Germany (calculated): {sample['Germany_Industrial'].iloc[0]:.6f}")
        
        # Export results
        export_industrial_demand(industrial_df)
        
        return industrial_df
        
    except Exception as e:
        logger.error(f"Error in main execution: {str(e)}")
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    industrial_data = main()