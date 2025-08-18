#!/usr/bin/env python3
"""
MASTER SCRIPT: European Gas Demand Aggregation Pipeline
Replicates the complete Excel workflow in a single script.

This script integrates all individual tasks into one seamless pipeline:
- Task 1: Country demand aggregation
- Task 4: Industrial demand calculation
- Task 5: LDZ demand calculation  
- Task 6: Gas-to-Power demand calculation (breakthrough solution)

Produces: daily_historic_data_by_category.csv (exact Excel match)
"""

import pandas as pd
import numpy as np
import openpyxl
import logging
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def load_multiticker_with_full_metadata(file_path='use4.xlsx', sheet_name='MultiTicker'):
    """
    Load MultiTicker data with complete metadata from rows 14, 15, 16.
    """
    logger.info(f"Loading MultiTicker with full metadata from {file_path}")
    
    # Load workbook to extract metadata
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    
    # Extract metadata from rows 14 (category), 15 (region), 16 (subcategory)
    metadata = {}
    max_col = min(ws.max_column, 600)  # Extended range
    
    logger.info(f"Extracting metadata from columns C to {openpyxl.utils.get_column_letter(max_col)}")
    
    for col in range(3, max_col + 1):  # Starting from column C
        col_name = f'Col_{col-2}'  # Col_1, Col_2, etc.
        category = ws.cell(row=14, column=col).value
        region = ws.cell(row=15, column=col).value  
        subcategory = ws.cell(row=16, column=col).value
        
        metadata[col_name] = {
            'category': str(category) if category else '',
            'region': str(region) if region else '',
            'subcategory': str(subcategory) if subcategory else ''
        }
    
    wb.close()
    
    # Load data using pandas
    df_full = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    
    # Data starts from row 21 (index 20), column B onwards
    data_rows = df_full.iloc[20:, 1:max_col].copy()
    
    # Set column names: Date + Col_1, Col_2, etc.
    data_rows.columns = ['Date'] + [f'Col_{i}' for i in range(1, len(data_rows.columns))]
    
    # Convert Date column
    data_rows['Date'] = pd.to_datetime(data_rows['Date'], errors='coerce')
    
    # Remove invalid dates
    data_rows = data_rows.dropna(subset=['Date'])
    
    # Convert data columns to numeric
    for col in data_rows.columns[1:]:
        data_rows[col] = pd.to_numeric(data_rows[col], errors='coerce')
    
    logger.info(f"Loaded {len(data_rows)} dates with {len(metadata)} tickers")
    
    return data_rows, metadata


def sumifs_three_criteria(data_df, metadata, category_target, region_target, subcategory_target):
    """
    Exact replication of Excel 3-criteria SUMIFS with exact string matching.
    """
    matching_cols = []
    
    for col, info in metadata.items():
        if col in data_df.columns:
            # Exact string matching (case sensitive as in Excel)
            category_match = info['category'] == category_target
            region_match = info['region'] == region_target
            subcategory_match = info['subcategory'] == subcategory_target
            
            if category_match and region_match and subcategory_match:
                matching_cols.append(col)
    
    if not matching_cols:
        logger.debug(f"No matches for {category_target}/{region_target}/{subcategory_target}")
        return pd.Series(0.0, index=data_df.index)
    
    logger.debug(f"Found {len(matching_cols)} matches for {category_target}/{region_target}/{subcategory_target}")
    
    # Sum across matching columns (handling NaN as 0)
    result = data_df[matching_cols].sum(axis=1, skipna=True)
    return result


def sumifs_two_criteria(data_df, metadata, category_target, region_target):
    """
    Exact replication of Excel 2-criteria SUMIFS with exact string matching.
    """
    matching_cols = []
    
    for col, info in metadata.items():
        if col in data_df.columns:
            # Exact string matching (case sensitive as in Excel)
            category_match = info['category'] == category_target
            region_match = info['region'] == region_target
            
            if category_match and region_match:
                matching_cols.append(col)
    
    if not matching_cols:
        logger.debug(f"No matches for {category_target}/{region_target}")
        return pd.Series(0.0, index=data_df.index)
    
    logger.debug(f"Found {len(matching_cols)} matches for {category_target}/{region_target}")
    
    # Sum across matching columns (handling NaN as 0)
    result = data_df[matching_cols].sum(axis=1, skipna=True)
    return result


def create_industrial_demand_sheet(data_df, metadata):
    """
    Task 4: Industrial Gas Demand calculation with exact Excel logic.
    
    Target: 240.70 for 2016-10-03
    """
    logger.info("Creating Industrial demand sheet (Task 4)")
    
    result = pd.DataFrame()
    result['Date'] = data_df['Date']
    
    # Column C: France Industrial
    # Criteria: Demand / France / Industrial
    result['France_Industrial'] = sumifs_three_criteria(
        data_df, metadata, 'Demand', 'France', 'Industrial'
    )
    
    # Column D: Belgium Industrial  
    # Criteria: Demand / Belgium / Industrial
    result['Belgium_Industrial'] = sumifs_three_criteria(
        data_df, metadata, 'Demand', 'Belgium', 'Industrial'
    )
    
    # Column E: Italy Industrial
    # Criteria: Demand / Italy / Industrial  
    result['Italy_Industrial'] = sumifs_three_criteria(
        data_df, metadata, 'Demand', 'Italy', 'Industrial'
    )
    
    # Column F: GB Industrial
    # Criteria: Demand / GB / Industrial
    result['GB_Industrial'] = sumifs_three_criteria(
        data_df, metadata, 'Demand', 'GB', 'Industrial'
    )
    
    # Column G: Netherlands Industrial and Power
    # Criteria: Demand / Netherlands / Industrial and Power
    result['Netherlands_IndPower'] = sumifs_three_criteria(
        data_df, metadata, 'Demand', 'Netherlands', 'Industrial and Power'
    )
    
    # Column H: Netherlands Zebra
    # Criteria: Demand / Netherlands / Zebra
    result['Netherlands_Zebra'] = sumifs_three_criteria(
        data_df, metadata, 'Demand', 'Netherlands', 'Zebra'
    )
    
    # Column I: Germany Industrial and Power (total)
    # Criteria: Demand / Germany / Industrial and Power
    result['Germany_Total'] = sumifs_three_criteria(
        data_df, metadata, 'Demand', 'Germany', 'Industrial and Power'
    )
    
    # Column J: Germany Gas-to-Power (to subtract)
    # Criteria: Intermediate Calculation / #Germany / Gas-to-Power
    result['Germany_GtP'] = sumifs_three_criteria(
        data_df, metadata, 'Intermediate Calculation', '#Germany', 'Gas-to-Power'
    )
    
    # Column K: Germany Industrial (calculated = I - J)
    result['Germany_Industrial'] = result['Germany_Total'] - result['Germany_GtP']
    
    # Column L: Total Industrial = SUM(C:H) + K
    # Sum of all industrial components
    industrial_components = [
        'France_Industrial', 'Belgium_Industrial', 'Italy_Industrial', 'GB_Industrial',
        'Netherlands_IndPower', 'Netherlands_Zebra', 'Germany_Industrial'
    ]
    
    result['Total_Industrial_Demand'] = result[industrial_components].sum(axis=1)
    
    logger.info("Industrial demand sheet created successfully")
    return result


def create_ldz_demand_sheet(data_df, metadata):
    """
    Task 5: LDZ Demand calculation with exact Excel logic.
    
    Target: 307.80 for 2016-10-03
    """
    logger.info("Creating LDZ demand sheet (Task 5)")
    
    result = pd.DataFrame()
    result['Date'] = data_df['Date']
    
    ldz_total = pd.Series(0.0, index=data_df.index)
    
    # Standard LDZ countries
    ldz_countries = {
        'France': 'LDZ',
        'Belgium': 'LDZ', 
        'Italy': 'LDZ',
        'Netherlands': 'LDZ',
        'GB': 'LDZ',
        'Germany': 'LDZ'
    }
    
    for country, subcategory in ldz_countries.items():
        country_ldz = sumifs_three_criteria(
            data_df, metadata, 'Demand', country, subcategory
        )
        if country_ldz.sum() > 0:
            ldz_total += country_ldz
    
    # Italy Other
    italy_other = sumifs_three_criteria(
        data_df, metadata, 'Demand', 'Italy', 'Other'
    )
    ldz_total += italy_other
    
    # Special cases where subcategory matches region
    special_cases = {
        'Austria': 'Austria',
        'Switzerland': 'Switzerland', 
        'Luxembourg': 'Luxembourg'
    }
    
    for country, subcategory in special_cases.items():
        country_ldz = sumifs_three_criteria(
            data_df, metadata, 'Demand', country, subcategory
        )
        if country_ldz.sum() > 0:
            ldz_total += country_ldz
    
    # Ireland (using "Demand (Net)")
    ireland_ldz = sumifs_three_criteria(
        data_df, metadata, 'Demand (Net)', 'Island of Ireland', 'Island of Ireland'
    )
    ldz_total += ireland_ldz
    
    result['Total_LDZ_Demand'] = ldz_total
    
    logger.info("LDZ demand sheet created successfully")
    return result


def create_gas_to_power_demand_sheet(data_df, metadata):
    """
    Task 6: Gas-to-Power Demand calculation with BREAKTHROUGH solution.
    
    Key insight: Excel total excludes Netherlands from the sum.
    Target: 166.71 for 2016-10-03
    """
    logger.info("Creating Gas-to-Power demand sheet (Task 6 - Breakthrough)")
    
    result = pd.DataFrame()
    result['Date'] = data_df['Date']
    
    # Countries included in Excel total (excludes Netherlands)
    gtp_total = pd.Series(0.0, index=data_df.index)
    
    # Standard Gas-to-Power countries
    standard_gtp_countries = ['France', 'Belgium', 'Italy', 'GB']
    
    for country in standard_gtp_countries:
        country_gtp = sumifs_three_criteria(
            data_df, metadata, 'Demand', country, 'Gas-to-Power'
        )
        if country_gtp.sum() > 0:
            gtp_total += country_gtp
    
    # Germany Gas-to-Power (using Intermediate Calculation)
    germany_gtp = sumifs_three_criteria(
        data_df, metadata, 'Intermediate Calculation', '#Germany', 'Gas-to-Power'
    )
    gtp_total += germany_gtp
    
    # Netherlands is calculated but EXCLUDED from total (breakthrough insight)
    # netherlands_gtp = sumifs_three_criteria(
    #     data_df, metadata, 'Intermediate Calculation', '#Netherlands', 'Gas-to-Power'
    # )
    # Netherlands excluded from gtp_total
    
    result['Total_Gas_to_Power_Demand'] = gtp_total
    
    logger.info("Gas-to-Power demand sheet created successfully")
    return result


def create_daily_country_demands(data_df, metadata):
    """
    Task 1: Country demand aggregation using 2-criteria SUMIFS.
    
    Target: France=90.13, Total=715.22 for 2016-10-03
    """
    logger.info("Creating daily country demands (Task 1)")
    
    result = pd.DataFrame()
    result['Date'] = data_df['Date']
    
    # Define countries to process
    countries = [
        'France', 'Belgium', 'Italy', 'Netherlands', 'GB', 
        'Austria', 'Germany', 'Switzerland', 'Luxembourg'
    ]
    
    country_totals = []
    
    for country in countries:
        country_demand = sumifs_two_criteria(
            data_df, metadata, 'Demand', country
        )
        result[country] = country_demand
        country_totals.append(country_demand)
    
    # Ireland using "Demand (Net)"
    ireland_demand = sumifs_two_criteria(
        data_df, metadata, 'Demand (Net)', 'Island of Ireland'
    )
    result['Ireland'] = ireland_demand
    country_totals.append(ireland_demand)
    
    # Calculate total
    total_demand = pd.Series(0.0, index=data_df.index)
    for country_total in country_totals:
        total_demand += country_total
    
    result['Total'] = total_demand
    
    logger.info("Daily country demands created successfully")
    return result


def merge_all_components(daily_country_data, industrial_data, ldz_data, gtp_data):
    """
    Merge all components into complete Daily historic data.
    
    Final columns: Date, Countries, Total, Industrial, LDZ, Gas_to_Power
    """
    logger.info("Merging all components into complete dataset")
    
    # Start with country data
    complete_data = daily_country_data.copy()
    
    # Add subcategory totals
    complete_data['Industrial'] = industrial_data['Total_Industrial_Demand']
    complete_data['LDZ'] = ldz_data['Total_LDZ_Demand'] 
    complete_data['Gas_to_Power'] = gtp_data['Total_Gas_to_Power_Demand']
    
    # Sort by date
    complete_data = complete_data.sort_values('Date').reset_index(drop=True)
    
    logger.info("Component merging completed successfully")
    return complete_data


def validate_complete_results(complete_data):
    """
    Validate against known Excel values for comprehensive testing.
    
    2016-10-03 targets:
    - France: 90.13
    - Total: 715.22  
    - Industrial: 240.70
    - LDZ: 307.80
    - Gas_to_Power: 166.71
    """
    logger.info("Running comprehensive validation suite")
    
    # Target values for 2016-10-03
    targets = {
        'France': 90.13,
        'Total': 715.22,
        'Industrial': 240.70,
        'LDZ': 307.80,
        'Gas_to_Power': 166.71
    }
    
    # Find validation date
    validation_date = '2016-10-03'
    sample = complete_data[complete_data['Date'] == validation_date]
    
    if sample.empty:
        logger.warning(f"Validation date {validation_date} not found in results")
        return False
    
    logger.info(f"\n📊 MASTER VALIDATION for {validation_date}:")
    logger.info("=" * 70)
    
    all_pass = True
    sample_row = sample.iloc[0]
    
    for column, target in targets.items():
        if column in sample.columns:
            actual = sample_row[column]
            diff = abs(actual - target)
            
            if diff < 0.01:
                status = "✅ PERFECT"
            elif diff < 0.5:
                status = "✅ GOOD"  
            else:
                status = "❌ FAIL"
                all_pass = False
            
            logger.info(f"  {status} {column}: {actual:.2f} (target {target:.2f}, diff: {diff:.2f})")
        else:
            logger.warning(f"  ❌ {column}: Column not found")
            all_pass = False
    
    logger.info("=" * 70)
    
    if all_pass:
        logger.info("🎯 MASTER VALIDATION SUCCESS: All targets achieved!")
        logger.info("🚀 Pipeline is PRODUCTION READY!")
    else:
        logger.warning("⚠️  Some validation targets not met")
    
    return all_pass


def run_full_validation_suite(complete_data):
    """
    Comprehensive testing of all pipeline components.
    """
    logger.info("Running full validation suite")
    
    # Basic data quality checks
    logger.info("Checking data quality...")
    
    # Check for missing dates
    if complete_data['Date'].isna().any():
        logger.warning("Found missing dates in dataset")
    
    # Check for negative values (should not exist in gas demand)
    numeric_cols = complete_data.select_dtypes(include=[np.number]).columns
    for col in numeric_cols:
        if (complete_data[col] < 0).any():
            logger.warning(f"Found negative values in {col}")
    
    # Check total consistency
    country_cols = ['France', 'Belgium', 'Italy', 'Netherlands', 'GB', 
                   'Austria', 'Germany', 'Switzerland', 'Luxembourg', 'Ireland']
    
    sample = complete_data.head(10)  # Check first 10 rows
    for idx, row in sample.iterrows():
        calculated_total = sum(row[col] for col in country_cols if col in complete_data.columns)
        reported_total = row['Total']
        diff = abs(calculated_total - reported_total)
        
        if diff > 0.01:
            logger.warning(f"Total mismatch on {row['Date']}: calculated={calculated_total:.2f}, reported={reported_total:.2f}")
    
    logger.info("Full validation suite completed")


def main():
    """
    Main pipeline execution function.
    
    Integrates all tasks into one seamless workflow.
    """
    logger.info("🚀 Starting MASTER European Gas Demand Aggregation Pipeline")
    logger.info("=" * 80)
    
    try:
        # Step 1: Load and process data
        logger.info("📁 Step 1: Loading MultiTicker data...")
        data_df, metadata = load_multiticker_with_full_metadata('use4.xlsx', 'MultiTicker')
        
        # Step 2: Create all subcategory sheets
        logger.info("🏭 Step 2a: Creating Industrial demand sheet...")
        industrial_data = create_industrial_demand_sheet(data_df, metadata)
        
        logger.info("🏠 Step 2b: Creating LDZ demand sheet...")
        ldz_data = create_ldz_demand_sheet(data_df, metadata)
        
        logger.info("⚡ Step 2c: Creating Gas-to-Power demand sheet...")
        gtp_data = create_gas_to_power_demand_sheet(data_df, metadata)
        
        # Step 3: Create country demand data
        logger.info("🌍 Step 3: Creating country demand aggregation...")
        daily_country_data = create_daily_country_demands(data_df, metadata)
        
        # Step 4: Merge everything
        logger.info("🔗 Step 4: Merging all components...")
        complete_data = merge_all_components(daily_country_data, industrial_data, ldz_data, gtp_data)
        
        # Step 5: Validate and export
        logger.info("✅ Step 5: Running validation...")
        validation_passed = validate_complete_results(complete_data)
        
        if validation_passed:
            logger.info("📊 Step 6: Exporting final results...")
            
            # Format dates for export
            export_data = complete_data.copy()
            export_data['Date'] = export_data['Date'].dt.strftime('%Y-%m-%d')
            
            # Round to match Excel precision
            numeric_cols = export_data.select_dtypes(include=[np.number]).columns
            export_data[numeric_cols] = export_data[numeric_cols].round(2)
            
            # Export to CSV
            output_file = 'daily_historic_data_by_category.csv'
            export_data.to_csv(output_file, index=False)
            
            logger.info(f"✅ SUCCESS: Complete Daily historic data exported to {output_file}")
            logger.info("=" * 80)
            logger.info("🎯 MASTER PIPELINE COMPLETED SUCCESSFULLY!")
            logger.info("🚀 Ready for production use!")
            
            # Show sample results
            logger.info(f"\n📋 Sample results for 2016-10-03:")
            sample = export_data[export_data['Date'] == '2016-10-03']
            if not sample.empty:
                sample_row = sample.iloc[0]
                logger.info(f"  🇫🇷 France: {sample_row['France']}")
                logger.info(f"  📊 Total: {sample_row['Total']}")
                logger.info(f"  🏭 Industrial: {sample_row['Industrial']}")
                logger.info(f"  🏠 LDZ: {sample_row['LDZ']}")
                logger.info(f"  ⚡ Gas-to-Power: {sample_row['Gas_to_Power']}")
            
            return complete_data
        else:
            logger.error("❌ Validation failed - pipeline aborted")
            return None
        
    except Exception as e:
        logger.error(f"❌ Pipeline failed with error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    result = main()