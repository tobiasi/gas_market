#!/usr/bin/env python3
"""
COMPLETE EUROPEAN GAS SUPPLY PROCESSOR

Comprehensive supply-side processing that extracts ALL 20 supply routes from MultiTicker data,
implementing the complete supply route breakdown found in Excel columns R-AM.

This module provides:
1. Complete supply route extraction (ALL 20 routes from Excel columns R-AM)
2. MultiTicker-based SUMIFS aggregation (NOT direct Excel extraction)  
3. Pipeline imports, LNG, domestic production, exports, and other flows
4. Geopolitical corrections (Russian Nord Stream post-2023)
5. Total supply calculation with ALL components
6. Clean 2017-01-01 start date alignment
7. Integration with existing demand-side processing

CRITICAL: Uses MultiTicker tab data with proper metadata structure (rows 14-16) for aggregation.
"""

import sys
sys.path.append("C:/development/commodities")

import pandas as pd
import numpy as np
import logging
from typing import Dict, List, Tuple, Optional
from datetime import datetime
import warnings

warnings.filterwarnings('ignore')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class CompleteSupplyProcessor:
    """
    Complete European gas supply processor for ALL 20 supply routes.
    
    Extracts comprehensive supply data from MultiTicker using SUMIFS logic
    to replicate Excel columns R-AM supply breakdown.
    """
    
    def __init__(self):
        """Initialize complete supply processor with all route mappings."""
        self.supply_audit_trail = []
        self.geopolitical_corrections = []
        
        # COMPLETE SUPPLY ROUTE MAPPING (Excel columns R-AM)
        # Based on MultiTicker analysis and Excel reference structure
        
        # PIPELINE IMPORTS (12 routes) - Excel columns R, S, T, X, Y, Z, AA, AB, AD, AE, AF, AL
        self.pipeline_imports = {
            'Slovakia_Austria': ('Import', 'Slovakia', 'Austria'),                    # Excel Column R
            'Russia_NordStream_Germany': ('Import', 'Russia (Nord Stream)', 'Germany'), # Excel Column S  
            'Norway_Europe': ('Import', 'Norway', 'Europe'),                         # Excel Column T
            'Algeria_Italy': ('Import', 'Algeria', 'Italy'),                         # Excel Column X
            'Libya_Italy': ('Import', 'Libya', 'Italy'),                            # Excel Column Y
            'Spain_France': ('Import', 'Spain', 'France'),                          # Excel Column Z
            'Denmark_Germany': ('Import', 'Denmark', 'Germany'),                    # Excel Column AA
            'Czech_Poland_Germany': ('Import', 'Czech and Poland', 'Germany'),       # Excel Column AB
            'Slovenia_Austria': ('Import', 'Slovenia', 'Austria'),                   # Excel Column AD
            'MAB_Austria': ('Import', 'MAB', 'Austria'),                            # Excel Column AE
            'TAP_Italy': ('Import', 'TAP', 'Italy'),                               # Excel Column AF
            'North_Africa_Imports': ('Import', 'North Africa', '*'),                # Excel Column AL
        }
        
        # LNG IMPORTS - Excel Column W (Total) + country breakdown
        self.lng_routes = {
            'LNG_Total': ('Import', 'LNG', '*'),                                    # Excel Column W
            'LNG_Belgium': ('Import', 'LNG', 'Belgium'),
            'LNG_France': ('Import', 'LNG', 'France'), 
            'LNG_Germany': ('Import', 'LNG', 'Germany'),
            'LNG_Netherlands': ('Import', 'LNG', 'Netherlands'),
            'LNG_GB': ('Import', 'LNG', 'GB'),
            'LNG_Italy': ('Import', 'LNG', 'Italy'),
        }
        
        # DOMESTIC PRODUCTION (6 sources) - Excel columns U, V, AG, AH, AI, AM
        self.domestic_production = {
            'Netherlands_Production': ('Production', 'Netherlands', 'Netherlands'),   # Excel Column U
            'GB_Production': ('Production', 'GB', 'GB'),                            # Excel Column V
            'Austria_Production': ('Production', 'Austria', 'Austria'),              # Excel Column AG
            'Italy_Production': ('Production', 'Italy', 'Italy'),                   # Excel Column AH
            'Germany_Production': ('Production', 'Germany', 'Germany'),              # Excel Column AI
            'Other_Production': ('Production', 'Other Europe', '*'),                 # Excel Column AM
        }
        
        # EXPORTS - Excel Column AC
        self.export_routes = {
            'Austria_Hungary_Export': ('Export', 'Austria', 'Hungary'),              # Excel Column AC
        }
        
        # OTHER FLOWS - Excel Column AK
        self.other_flows = {
            'Other_Border_Net_Flows': ('Supply', 'Border Points', '*'),             # Excel Column AK
        }
        
        # COMPLETE ROUTE REGISTRY (ALL 20+ routes)
        self.all_supply_routes = {
            **self.pipeline_imports,
            **self.lng_routes, 
            **self.domestic_production,
            **self.export_routes,
            **self.other_flows
        }
    
    def load_multiticker_supply_data(self, file_path: str = 'use4.xlsx', sheet_name: str = 'MultiTicker') -> Tuple[pd.DataFrame, Dict]:
        """
        Load MultiTicker data with focus on supply-side extraction.
        
        Returns MultiTicker data with proper metadata for supply route aggregation.
        """
        logger.info(f"📊 Loading MultiTicker supply data from {file_path}")
        
        # Load workbook to extract metadata
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        
        # Extract metadata from rows 14 (category), 15 (region_from), 16 (region_to)
        metadata = {}
        max_col = min(ws.max_column, 600)
        
        logger.info(f"Extracting supply metadata from columns C to {openpyxl.utils.get_column_letter(max_col)}")
        
        for col in range(3, max_col + 1):
            col_name = f'Col_{col-2}'
            category = ws.cell(row=14, column=col).value
            region_from = ws.cell(row=15, column=col).value
            region_to = ws.cell(row=16, column=col).value
            
            metadata[col_name] = {
                'category': str(category) if category else '',
                'region_from': str(region_from) if region_from else '',  # Row 15
                'region_to': str(region_to) if region_to else ''         # Row 16  
            }
        
        wb.close()
        
        # Load data using pandas  
        df_full = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        
        # Data starts from row 21 (index 20), column B onwards
        data_rows = df_full.iloc[20:, 1:max_col].copy()
        
        # Set column names: Date + Col_1, Col_2, etc.
        data_rows.columns = ['Date'] + [f'Col_{i}' for i in range(1, len(data_rows.columns))]
        
        # Convert Date column
        data_rows['Date'] = pd.to_datetime(data_rows['Date'], errors='coerce')
        
        # Remove invalid dates and apply 2017-01-01 start filter
        data_rows = data_rows.dropna(subset=['Date'])
        data_rows = data_rows[data_rows['Date'] >= '2017-01-01'].copy()  # Clean start date
        
        # Convert data columns to numeric
        for col in data_rows.columns[1:]:
            data_rows[col] = pd.to_numeric(data_rows[col], errors='coerce')
        
        logger.info(f"Loaded {len(data_rows)} dates (from 2017-01-01) with {len(metadata)} tickers")
        
        return data_rows, metadata
    
    def sumifs_supply_route(self, data_df: pd.DataFrame, metadata: Dict,
                           category: str, region_from: str, region_to: str) -> pd.Series:
        """
        SUMIFS aggregation for supply routes using MultiTicker metadata structure.
        
        Implements 3-criteria SUMIFS:
        - Category (Row 14): Import/Export/Production/Supply
        - Region From (Row 15): Source country/region
        - Region To (Row 16): Destination country/region (* for aggregated)
        """
        matching_cols = []
        
        for col, info in metadata.items():
            if col in data_df.columns:
                # 3-criteria exact matching
                category_match = info['category'] == category
                region_from_match = info['region_from'] == region_from or region_from == '*'
                region_to_match = info['region_to'] == region_to or region_to == '*'
                
                if category_match and region_from_match and region_to_match:
                    matching_cols.append(col)
        
        if not matching_cols:
            logger.debug(f"No matches for {category}/{region_from}/{region_to}")
            return pd.Series(0.0, index=data_df.index)
        
        logger.debug(f"Found {len(matching_cols)} matches for {category}/{region_from}/{region_to}")
        
        # Sum across matching columns
        result = data_df[matching_cols].sum(axis=1, skipna=True)
        return result
    
    def process_pipeline_imports(self, data_df: pd.DataFrame, metadata: Dict) -> pd.DataFrame:
        """
        Process all 12 pipeline import routes (Excel columns R, S, T, X, Y, Z, AA, AB, AD, AE, AF, AL).
        """
        logger.info("🛢️ Processing all 12 pipeline import routes")
        
        result = pd.DataFrame()
        result['Date'] = data_df['Date']
        
        # Process each pipeline import route
        for route_name, (category, region_from, region_to) in self.pipeline_imports.items():
            route_data = self.sumifs_supply_route(data_df, metadata, category, region_from, region_to)
            result[route_name] = route_data
            
            if route_data.sum() != 0:  # Log non-zero routes
                logger.debug(f"✅ {route_name}: {route_data.sum():.2f} total flow")
        
        # Calculate total pipeline imports
        pipeline_cols = list(self.pipeline_imports.keys())
        result['Total_Pipeline_Imports'] = result[pipeline_cols].sum(axis=1)
        
        logger.info(f"✅ Pipeline imports processed: {len(self.pipeline_imports)} routes")
        
        # Audit trail
        self.supply_audit_trail.append({
            'processing_type': 'pipeline_imports',
            'routes_processed': len(self.pipeline_imports),
            'total_routes': 12,
            'excel_columns': 'R,S,T,X,Y,Z,AA,AB,AD,AE,AF,AL',
            'reason': 'Complete pipeline import extraction from MultiTicker'
        })
        
        return result
    
    def process_lng_imports(self, data_df: pd.DataFrame, metadata: Dict) -> pd.DataFrame:
        """
        Process LNG imports with total aggregation and country breakdown (Excel column W + details).
        """
        logger.info("🚢 Processing LNG imports with total and country breakdown")
        
        result = pd.DataFrame()
        result['Date'] = data_df['Date']
        
        # Process LNG routes
        for route_name, (category, region_from, region_to) in self.lng_routes.items():
            lng_data = self.sumifs_supply_route(data_df, metadata, category, region_from, region_to)
            result[route_name] = lng_data
            
            if lng_data.sum() != 0:
                logger.debug(f"✅ {route_name}: {lng_data.sum():.2f} total flow")
        
        logger.info(f"✅ LNG imports processed: {len(self.lng_routes)} routes")
        
        # Audit trail
        self.supply_audit_trail.append({
            'processing_type': 'lng_imports',
            'routes_processed': len(self.lng_routes),
            'excel_column': 'W',
            'reason': 'LNG total + country breakdown from MultiTicker'
        })
        
        return result
    
    def process_domestic_production(self, data_df: pd.DataFrame, metadata: Dict) -> pd.DataFrame:
        """
        Process all 6 domestic production sources (Excel columns U, V, AG, AH, AI, AM).
        """
        logger.info("🏭 Processing all 6 domestic production sources")
        
        result = pd.DataFrame()
        result['Date'] = data_df['Date']
        
        # Process each production source
        for route_name, (category, region_from, region_to) in self.domestic_production.items():
            production_data = self.sumifs_supply_route(data_df, metadata, category, region_from, region_to)
            result[route_name] = production_data
            
            if production_data.sum() != 0:
                logger.debug(f"✅ {route_name}: {production_data.sum():.2f} total production")
        
        # Calculate total domestic production
        production_cols = list(self.domestic_production.keys())
        result['Total_Domestic_Production'] = result[production_cols].sum(axis=1)
        
        logger.info(f"✅ Domestic production processed: {len(self.domestic_production)} sources")
        
        # Audit trail
        self.supply_audit_trail.append({
            'processing_type': 'domestic_production',
            'routes_processed': len(self.domestic_production),
            'total_sources': 6,
            'excel_columns': 'U,V,AG,AH,AI,AM',
            'reason': 'Complete domestic production extraction from MultiTicker'
        })
        
        return result
    
    def process_exports_and_other_flows(self, data_df: pd.DataFrame, metadata: Dict) -> pd.DataFrame:
        """
        Process exports (Excel column AC) and other flows (Excel column AK).
        """
        logger.info("📤 Processing exports and other border flows")
        
        result = pd.DataFrame()
        result['Date'] = data_df['Date']
        
        # Process export routes
        for route_name, (category, region_from, region_to) in self.export_routes.items():
            export_data = self.sumifs_supply_route(data_df, metadata, category, region_from, region_to)
            result[route_name] = export_data
            
            if export_data.sum() != 0:
                logger.debug(f"✅ {route_name}: {export_data.sum():.2f} total export")
        
        # Process other flows
        for route_name, (category, region_from, region_to) in self.other_flows.items():
            flow_data = self.sumifs_supply_route(data_df, metadata, category, region_from, region_to)
            result[route_name] = flow_data
            
            if flow_data.sum() != 0:
                logger.debug(f"✅ {route_name}: {flow_data.sum():.2f} total flow")
        
        logger.info("✅ Exports and other flows processed")
        
        # Audit trail
        self.supply_audit_trail.append({
            'processing_type': 'exports_other_flows',
            'routes_processed': len(self.export_routes) + len(self.other_flows),
            'excel_columns': 'AC,AK',
            'reason': 'Export flows and other border flows from MultiTicker'
        })
        
        return result
    
    def apply_geopolitical_corrections(self, supply_data: pd.DataFrame) -> pd.DataFrame:
        """
        Apply geopolitical corrections to supply data.
        
        Key corrections:
        - Russian Nord Stream supply zeroed post-2023-01-01
        """
        logger.info("🌍 Applying geopolitical corrections")
        
        corrected_data = supply_data.copy()
        corrections_applied = 0
        
        if 'Russia_NordStream_Germany' in corrected_data.columns:
            # Get date-based mask for post-2023 corrections
            dates = pd.to_datetime(corrected_data['Date'])
            post_2023_mask = dates >= '2023-01-01'
            
            if post_2023_mask.any():
                original_sum = corrected_data.loc[post_2023_mask, 'Russia_NordStream_Germany'].sum()
                corrected_data.loc[post_2023_mask, 'Russia_NordStream_Germany'] = 0.0
                corrections_applied = post_2023_mask.sum()
                
                logger.info(f"✅ Russian Nord Stream zeroed: {corrections_applied} dates post-2023-01-01")
                
                # Audit trail
                self.geopolitical_corrections.append({
                    'route': 'Russia_NordStream_Germany',
                    'correction_type': 'geopolitical_zeroing',
                    'date_from': '2023-01-01',
                    'dates_affected': corrections_applied,
                    'original_sum': original_sum,
                    'corrected_sum': 0.0,
                    'reason': 'Post-conflict Russian supply disruption'
                })
        
        logger.info(f"✅ Geopolitical corrections completed: {len(self.geopolitical_corrections)} corrections")
        
        return corrected_data
    
    def calculate_total_supply(self, pipeline_data: pd.DataFrame, lng_data: pd.DataFrame,
                             production_data: pd.DataFrame, exports_other_data: pd.DataFrame) -> pd.Series:
        """
        Calculate Total_Supply using ALL components (Excel column AJ equivalent).
        
        Total_Supply = Pipeline Imports + LNG + Domestic Production + Exports + Other Flows
        """
        logger.info("📊 Calculating Total Supply (ALL components)")
        
        total_supply = pd.Series(0.0, index=pipeline_data.index)
        
        # Add pipeline imports
        if 'Total_Pipeline_Imports' in pipeline_data.columns:
            total_supply += pipeline_data['Total_Pipeline_Imports']
        
        # Add LNG total
        if 'LNG_Total' in lng_data.columns:
            total_supply += lng_data['LNG_Total']
        
        # Add domestic production
        if 'Total_Domestic_Production' in production_data.columns:
            total_supply += production_data['Total_Domestic_Production']
        
        # Add exports (typically negative)
        for export_col in self.export_routes.keys():
            if export_col in exports_other_data.columns:
                total_supply += exports_other_data[export_col]
        
        # Add other flows
        for other_col in self.other_flows.keys():
            if other_col in exports_other_data.columns:
                total_supply += exports_other_data[other_col]
        
        logger.info(f"✅ Total Supply calculated from ALL components")
        
        return total_supply
    
    def process_complete_supply_side(self, file_path: str = 'use4.xlsx') -> pd.DataFrame:
        """
        Process complete supply-side data with ALL 20 routes.
        
        Returns comprehensive supply data DataFrame with all routes from Excel columns R-AM.
        """
        logger.info("🚀 Starting COMPLETE supply-side processing")
        logger.info("=" * 80)
        logger.info("EXTRACTING ALL 20 SUPPLY ROUTES (Excel columns R-AM)")
        logger.info("=" * 80)
        
        try:
            # Step 1: Load MultiTicker supply data
            logger.info("📊 Step 1: Loading MultiTicker supply data...")
            data_df, metadata = self.load_multiticker_supply_data(file_path)
            
            # Step 2: Process pipeline imports (12 routes)
            logger.info("🛢️ Step 2: Processing pipeline imports...")
            pipeline_data = self.process_pipeline_imports(data_df, metadata)
            
            # Step 3: Process LNG imports  
            logger.info("🚢 Step 3: Processing LNG imports...")
            lng_data = self.process_lng_imports(data_df, metadata)
            
            # Step 4: Process domestic production (6 sources)
            logger.info("🏭 Step 4: Processing domestic production...")
            production_data = self.process_domestic_production(data_df, metadata)
            
            # Step 5: Process exports and other flows
            logger.info("📤 Step 5: Processing exports and other flows...")
            exports_other_data = self.process_exports_and_other_flows(data_df, metadata)
            
            # Step 6: Combine all supply data
            logger.info("🔗 Step 6: Combining all supply data...")
            complete_supply = pipeline_data.copy()
            
            # Add LNG data
            for col in lng_data.columns:
                if col != 'Date':
                    complete_supply[col] = lng_data[col]
            
            # Add production data  
            for col in production_data.columns:
                if col != 'Date':
                    complete_supply[col] = production_data[col]
            
            # Add exports and other flows
            for col in exports_other_data.columns:
                if col != 'Date':
                    complete_supply[col] = exports_other_data[col]
            
            # Step 7: Calculate Total Supply
            logger.info("📊 Step 7: Calculating Total Supply...")
            complete_supply['Total_Supply'] = self.calculate_total_supply(
                pipeline_data, lng_data, production_data, exports_other_data
            )
            
            # Step 8: Apply geopolitical corrections
            logger.info("🌍 Step 8: Applying geopolitical corrections...")
            complete_supply = self.apply_geopolitical_corrections(complete_supply)
            
            # Recalculate totals after corrections
            pipeline_cols = list(self.pipeline_imports.keys())
            production_cols = list(self.domestic_production.keys())
            
            complete_supply['Total_Pipeline_Imports'] = complete_supply[pipeline_cols].sum(axis=1)
            complete_supply['Total_Domestic_Production'] = complete_supply[production_cols].sum(axis=1)
            
            # Recalculate Total Supply after corrections
            complete_supply['Total_Supply'] = self.calculate_total_supply(
                complete_supply, complete_supply, complete_supply, complete_supply
            )
            
            logger.info("=" * 80)
            logger.info("✅ COMPLETE SUPPLY-SIDE PROCESSING FINISHED")
            logger.info(f"📊 Processed {len(self.all_supply_routes)} total supply routes")
            logger.info(f"📅 Date range: {complete_supply['Date'].min()} to {complete_supply['Date'].max()}")
            logger.info(f"📈 Data points: {len(complete_supply)} dates")
            
            return complete_supply
            
        except Exception as e:
            logger.error(f"❌ Complete supply processing failed: {str(e)}")
            import traceback
            traceback.print_exc()
            raise
    
    def export_supply_audit_trail(self, output_file: str = 'complete_supply_audit_trail.csv') -> str:
        """Export complete supply processing audit trail."""
        if not self.supply_audit_trail:
            logger.warning("No supply audit trail to export")
            return None
        
        audit_df = pd.DataFrame(self.supply_audit_trail)
        audit_df.to_csv(output_file, index=False)
        
        logger.info(f"📝 Complete supply audit trail exported to {output_file}")
        return output_file


def main():
    """
    Demonstration of complete supply processor.
    """
    logger.info("🚀 COMPLETE SUPPLY PROCESSOR DEMO")
    logger.info("=" * 80)
    
    try:
        # Initialize processor
        processor = CompleteSupplyProcessor()
        
        # Show all route mappings
        logger.info("🛢️ PIPELINE IMPORTS (12 routes):")
        for route, (cat, from_loc, to_loc) in processor.pipeline_imports.items():
            logger.info(f"  {route}: {cat} {from_loc} → {to_loc}")
        
        logger.info("\n🚢 LNG IMPORTS:")
        for route, (cat, from_loc, to_loc) in processor.lng_routes.items():
            logger.info(f"  {route}: {cat} {from_loc} → {to_loc}")
        
        logger.info("\n🏭 DOMESTIC PRODUCTION (6 sources):")
        for route, (cat, from_loc, to_loc) in processor.domestic_production.items():
            logger.info(f"  {route}: {cat} {from_loc} → {to_loc}")
        
        logger.info("\n📤 EXPORTS & OTHER FLOWS:")
        for route, (cat, from_loc, to_loc) in {**processor.export_routes, **processor.other_flows}.items():
            logger.info(f"  {route}: {cat} {from_loc} → {to_loc}")
        
        logger.info(f"\n📊 TOTAL ROUTES: {len(processor.all_supply_routes)}")
        logger.info("=" * 80)
        logger.info("✅ Complete Supply Processor Ready")
        logger.info("🔗 Ready for integration with demand-side pipeline")
        
        return processor
        
    except Exception as e:
        logger.error(f"❌ Error in complete supply demo: {str(e)}")
        raise


if __name__ == "__main__":
    processor = main()