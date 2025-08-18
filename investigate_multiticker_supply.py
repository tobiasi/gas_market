#!/usr/bin/env python3
"""
MULTITICKER SUPPLY ROUTE INVESTIGATION

This script investigates the actual MultiTicker structure to understand
what supply-related data is available and how it's categorized.
"""

import sys
sys.path.append("C:/development/commodities")

import pandas as pd
import numpy as np
import openpyxl
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def investigate_multiticker_supply_structure(file_path='use4.xlsx', sheet_name='MultiTicker'):
    """
    Investigate the MultiTicker structure for supply-related data.
    """
    logger.info("🔍 INVESTIGATING MULTITICKER SUPPLY STRUCTURE")
    logger.info("=" * 80)
    
    try:
        # Load workbook to extract metadata
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        
        # Extract metadata from rows 14 (category), 15 (region), 16 (subcategory)
        logger.info("📊 Extracting supply-relevant metadata...")
        
        supply_metadata = {}
        max_col = min(ws.max_column, 600)
        
        for col in range(3, max_col + 1):
            col_name = f'Col_{col-2}'
            category = ws.cell(row=14, column=col).value
            region = ws.cell(row=15, column=col).value
            subcategory = ws.cell(row=16, column=col).value
            
            # Look for supply-related categories
            if category and any(supply_keyword in str(category).lower() for supply_keyword in 
                              ['import', 'export', 'production', 'supply', 'lng']):
                supply_metadata[col_name] = {
                    'category': str(category) if category else '',
                    'region': str(region) if region else '',
                    'subcategory': str(subcategory) if subcategory else '',
                    'excel_col': openpyxl.utils.get_column_letter(col)
                }
        
        wb.close()
        
        logger.info(f"✅ Found {len(supply_metadata)} supply-related columns")
        
        # Group by category
        by_category = {}
        for col_name, info in supply_metadata.items():
            category = info['category']
            if category not in by_category:
                by_category[category] = []
            by_category[category].append((col_name, info))
        
        logger.info("\\n📋 SUPPLY CATEGORIES FOUND:")
        for category, columns in by_category.items():
            logger.info(f"\\n🔶 {category} ({len(columns)} columns):")
            for col_name, info in columns[:20]:  # Show first 20 of each category
                logger.info(f"   {col_name} ({info['excel_col']}) | {info['region']} → {info['subcategory']}")
            if len(columns) > 20:
                logger.info(f"   ... and {len(columns)-20} more columns")
        
        # Look for specific routes we're missing
        logger.info("\\n🎯 SEARCHING FOR MISSING ROUTES:")
        missing_routes = ['GB_Production', 'LNG_Total', 'Spain_France', 'Denmark_Germany', 
                         'Czech_Poland_Germany', 'Austria_Hungary_Export', 'Other_Border_Net_Flows']
        
        for route in missing_routes:
            logger.info(f"\\n🔍 Searching for {route}:")
            
            # Search by keywords
            keywords = route.lower().replace('_', ' ').split()
            potential_matches = []
            
            for col_name, info in supply_metadata.items():
                region_lower = info['region'].lower()
                subcategory_lower = info['subcategory'].lower()
                category_lower = info['category'].lower()
                
                match_score = 0
                for keyword in keywords:
                    if keyword in region_lower or keyword in subcategory_lower or keyword in category_lower:
                        match_score += 1
                
                if match_score > 0:
                    potential_matches.append((col_name, info, match_score))
            
            # Sort by match score
            potential_matches.sort(key=lambda x: x[2], reverse=True)
            
            if potential_matches:
                logger.info(f"   Potential matches for {route}:")
                for col_name, info, score in potential_matches[:5]:
                    logger.info(f"     {col_name} ({info['excel_col']}) | Score: {score} | {info['category']} | {info['region']} → {info['subcategory']}")
            else:
                logger.info(f"   ❌ No potential matches found for {route}")
        
        # Look for production data specifically
        logger.info("\\n🏭 PRODUCTION DATA ANALYSIS:")
        production_columns = []
        for col_name, info in supply_metadata.items():
            if 'production' in info['category'].lower() or 'production' in info['subcategory'].lower():
                production_columns.append((col_name, info))
        
        if production_columns:
            logger.info(f"Found {len(production_columns)} production columns:")
            for col_name, info in production_columns:
                logger.info(f"   {col_name} ({info['excel_col']}) | {info['region']} → {info['subcategory']}")
        else:
            logger.info("❌ No production columns found")
        
        # Look for LNG data specifically
        logger.info("\\n🚢 LNG DATA ANALYSIS:")
        lng_columns = []
        for col_name, info in supply_metadata.items():
            if 'lng' in info['category'].lower() or 'lng' in info['region'].lower() or 'lng' in info['subcategory'].lower():
                lng_columns.append((col_name, info))
        
        if lng_columns:
            logger.info(f"Found {len(lng_columns)} LNG columns:")
            for col_name, info in lng_columns:
                logger.info(f"   {col_name} ({info['excel_col']}) | {info['region']} → {info['subcategory']}")
        else:
            logger.info("❌ No LNG columns found")
        
        return supply_metadata, by_category
        
    except Exception as e:
        logger.error(f"❌ Investigation failed: {str(e)}")
        raise


def investigate_all_categories():
    """
    Investigate ALL categories in MultiTicker to understand the full structure.
    """
    logger.info("\\n🔍 INVESTIGATING ALL MULTITICKER CATEGORIES")
    logger.info("=" * 80)
    
    try:
        wb = openpyxl.load_workbook('use4.xlsx', read_only=True, data_only=True)
        ws = wb['MultiTicker']
        
        all_categories = set()
        all_regions = set()
        all_subcategories = set()
        
        max_col = min(ws.max_column, 600)
        
        for col in range(3, max_col + 1):
            category = ws.cell(row=14, column=col).value
            region = ws.cell(row=15, column=col).value
            subcategory = ws.cell(row=16, column=col).value
            
            if category:
                all_categories.add(str(category))
            if region:
                all_regions.add(str(region))
            if subcategory:
                all_subcategories.add(str(subcategory))
        
        wb.close()
        
        logger.info(f"📊 COMPLETE MULTITICKER STRUCTURE:")
        logger.info(f"\\n🔶 ALL CATEGORIES ({len(all_categories)}):")
        for category in sorted(all_categories):
            logger.info(f"   {category}")
        
        logger.info(f"\\n🌍 ALL REGIONS ({len(all_regions)}):")
        for region in sorted(all_regions):
            logger.info(f"   {region}")
        
        logger.info(f"\\n📋 ALL SUBCATEGORIES ({len(all_subcategories)}):")
        for subcategory in sorted(all_subcategories):
            logger.info(f"   {subcategory}")
            
    except Exception as e:
        logger.error(f"❌ Category investigation failed: {str(e)}")
        raise


def main():
    """
    Main investigation execution.
    """
    try:
        logger.info("🔍 MULTITICKER SUPPLY ROUTE INVESTIGATION")
        logger.info("=" * 80)
        
        # Investigate supply structure
        supply_metadata, by_category = investigate_multiticker_supply_structure()
        
        # Investigate all categories
        investigate_all_categories()
        
        logger.info("\\n✅ INVESTIGATION COMPLETE")
        logger.info("Use this information to correct the supply route mappings")
        
        return supply_metadata, by_category
        
    except Exception as e:
        logger.error(f"❌ Main investigation failed: {str(e)}")
        raise


if __name__ == "__main__":
    supply_metadata, by_category = main()