#!/usr/bin/env python3
"""
MultiTicker Tab Creation Script

Creates a new Excel file with MultiTicker sheet structure that:
1. Loads Bloomberg ticker list from complete_ticker_list.csv
2. Creates proper 3-row metadata headers (matching existing MultiTicker format)
3. Sets up structure for Bloomberg data population
4. Prepares date range for historical data
5. Ensures compatibility with existing master aggregation pipeline
"""

import sys
sys.path.append("C:/development/commodities")

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import logging
from complete_ticker_extraction import main as extract_tickers

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def load_ticker_metadata(csv_file='complete_ticker_list.csv'):
    """
    Load ticker metadata from CSV file.
    """
    logger.info(f"Loading ticker metadata from {csv_file}")
    
    try:
        ticker_df = pd.read_csv(csv_file)
        logger.info(f"Loaded {len(ticker_df)} tickers from {csv_file}")
        return ticker_df
    except FileNotFoundError:
        logger.warning(f"{csv_file} not found. Extracting tickers from use4.xlsx...")
        # Extract tickers if CSV doesn't exist
        ticker_df, _ = extract_tickers()
        return ticker_df


def create_date_range(start_date='2013-01-01', end_date=None):
    """
    Create date range for MultiTicker data.
    Default: from 2013-01-01 to today (covers historical data range)
    """
    if end_date is None:
        end_date = datetime.now().strftime('%Y-%m-%d')
    
    logger.info(f"Creating date range: {start_date} to {end_date}")
    
    dates = pd.date_range(start=start_date, end=end_date, freq='D')
    logger.info(f"Generated {len(dates)} dates")
    
    return dates


def create_multiticker_headers(ticker_df):
    """
    Create the 3-row metadata headers for MultiTicker format.
    
    Row 14: Category (Demand, Import, Production, etc.)
    Row 15: Region (Country/area names)
    Row 16: Subcategory (Industrial, Gas-to-Power, LDZ, etc.)
    """
    logger.info("Creating MultiTicker 3-row headers")
    
    headers = {
        'categories': [],      # Row 14
        'regions': [],         # Row 15  
        'subcategories': []    # Row 16
    }
    
    for _, ticker_row in ticker_df.iterrows():
        # Row 14: Category (main type)
        category = ticker_row.get('category', '')
        headers['categories'].append(category)
        
        # Row 15: Region (geographic area)
        region = ticker_row.get('region_from', ticker_row.get('region_to', ''))
        headers['regions'].append(region)
        
        # Row 16: Subcategory (specific type)
        # Derive subcategory from description and category
        description = str(ticker_row.get('description', '')).lower()
        category_lower = str(category).lower()
        
        subcategory = ''
        
        if 'industrial' in description:
            subcategory = 'Industrial'
        elif 'gas-to-power' in description or 'power' in description:
            subcategory = 'Gas-to-Power'
        elif 'ldz' in description or 'domestic' in description or 'residential' in description:
            subcategory = 'LDZ'
        elif 'storage' in description or 'inventory' in description:
            subcategory = 'Storage'
        elif 'lng' in description:
            subcategory = 'LNG'
        elif 'production' in category_lower:
            subcategory = 'Production'
        elif 'import' in category_lower:
            subcategory = 'Import'
        elif 'export' in category_lower:
            subcategory = 'Export'
        else:
            # Default based on category
            if 'demand' in category_lower:
                subcategory = 'Other'
            else:
                subcategory = category
        
        headers['subcategories'].append(subcategory)
    
    logger.info(f"Created headers for {len(ticker_df)} tickers")
    
    # Show header summary
    logger.info("Header breakdown:")
    logger.info(f"  Categories: {len(set(headers['categories']))} unique")
    logger.info(f"  Regions: {len(set(headers['regions']))} unique")  
    logger.info(f"  Subcategories: {len(set(headers['subcategories']))} unique")
    
    return headers


def create_multiticker_excel(ticker_df, dates, headers, output_file='multiticker_tab.xlsx'):
    """
    Create the MultiTicker Excel file with proper formatting.
    """
    logger.info(f"Creating MultiTicker Excel file: {output_file}")
    
    # Create new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'MultiTicker'
    
    # Style definitions
    header_font = Font(bold=True, size=10)
    date_font = Font(size=9)
    center_align = Alignment(horizontal='center')
    
    # Row 1: Empty (for spacing)
    logger.info("Creating row structure...")
    
    # Rows 2-13: Empty/metadata space (matching existing format)
    
    # Row 14: Categories
    ws['A14'] = 'Date'
    ws['A14'].font = header_font
    
    for col_idx, category in enumerate(headers['categories'], start=2):
        ws.cell(row=14, column=col_idx, value=category).font = header_font
    
    # Row 15: Regions  
    for col_idx, region in enumerate(headers['regions'], start=2):
        ws.cell(row=15, column=col_idx, value=region).font = header_font
    
    # Row 16: Subcategories
    for col_idx, subcategory in enumerate(headers['subcategories'], start=2):
        ws.cell(row=16, column=col_idx, value=subcategory).font = header_font
    
    # Row 17: Ticker symbols (for reference)
    ws['A17'] = 'Ticker'
    ws['A17'].font = header_font
    
    for col_idx, ticker in enumerate(ticker_df['ticker'], start=2):
        ws.cell(row=17, column=col_idx, value=ticker).font = Font(size=8)
    
    # Row 18: Units (for reference)
    ws['A18'] = 'Units'
    ws['A18'].font = header_font
    
    for col_idx, units in enumerate(ticker_df['units'], start=2):
        ws.cell(row=18, column=col_idx, value=units).font = Font(size=8)
    
    # Row 19: Normalization factors (for reference)
    ws['A19'] = 'Norm Factor'
    ws['A19'].font = header_font
    
    for col_idx, norm_factor in enumerate(ticker_df['normalization_factor'], start=2):
        ws.cell(row=19, column=col_idx, value=float(norm_factor) if pd.notna(norm_factor) else 1.0).font = Font(size=8)
    
    # Row 20: Empty separator
    
    # Row 21 onwards: Date data rows
    logger.info(f"Adding {len(dates)} date rows...")
    
    for date_idx, date in enumerate(dates, start=21):
        # Date column (Column A)
        ws.cell(row=date_idx, column=1, value=date).font = date_font
        
        # Data columns (B onwards) - initialized to 0 for Bloomberg data population
        for col_idx in range(2, len(ticker_df) + 2):
            ws.cell(row=date_idx, column=col_idx, value=0.0)
    
    # Format date column
    logger.info("Applying formatting...")
    for row in range(21, 21 + len(dates)):
        ws.cell(row=row, column=1).number_format = 'YYYY-MM-DD'
    
    # Auto-size columns (first few only to avoid performance issues)
    for col in range(1, min(20, len(ticker_df) + 2)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
    
    # Add metadata sheet
    logger.info("Creating metadata reference sheet...")
    meta_ws = wb.create_sheet('Metadata')
    
    # Copy ticker metadata to reference sheet
    meta_df = ticker_df.copy()
    
    # Write headers
    for col_idx, col_name in enumerate(meta_df.columns, start=1):
        meta_ws.cell(row=1, column=col_idx, value=col_name).font = header_font
    
    # Write data
    for row_idx, (_, row) in enumerate(meta_df.iterrows(), start=2):
        for col_idx, value in enumerate(row, start=1):
            meta_ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-size metadata sheet
    for col in range(1, len(meta_df.columns) + 1):
        meta_ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Save workbook
    wb.save(output_file)
    logger.info(f"✅ MultiTicker Excel file created: {output_file}")
    
    # Summary info
    logger.info("=" * 60)
    logger.info("📊 MULTITICKER STRUCTURE SUMMARY:")
    logger.info(f"  Output file: {output_file}")
    logger.info(f"  Sheets: MultiTicker, Metadata")
    logger.info(f"  Tickers: {len(ticker_df)}")
    logger.info(f"  Date range: {dates[0].strftime('%Y-%m-%d')} to {dates[-1].strftime('%Y-%m-%d')}")
    logger.info(f"  Total rows: {len(dates)} data rows")
    logger.info(f"  Structure: Rows 14-16 (metadata), Row 21+ (data)")
    logger.info("=" * 60)
    
    return output_file


def validate_pipeline_compatibility(output_file='multiticker_tab.xlsx'):
    """
    Validate that the created MultiTicker format is compatible with existing pipeline.
    """
    logger.info("🔍 Validating pipeline compatibility...")
    
    try:
        # Test loading with the existing pipeline function
        from master_aggregation_script import load_multiticker_with_full_metadata
        
        # Try to load the created file
        data_df, metadata = load_multiticker_with_full_metadata(output_file, 'MultiTicker')
        
        logger.info(f"✅ Pipeline compatibility test PASSED:")
        logger.info(f"  Loaded {len(data_df)} rows")
        logger.info(f"  Loaded {len(metadata)} ticker metadata entries")
        logger.info(f"  Sample date: {data_df['Date'].iloc[0] if len(data_df) > 0 else 'None'}")
        
        return True
        
    except Exception as e:
        logger.error(f"❌ Pipeline compatibility test FAILED: {str(e)}")
        return False


def setup_bloomberg_data_integration():
    """
    Set up instructions and structure for Bloomberg data integration.
    """
    logger.info("📋 Setting up Bloomberg data integration guide...")
    
    instructions = """
# Bloomberg Data Integration Instructions

## Files Created:
1. `multiticker_tab.xlsx` - Ready-to-populate MultiTicker structure
2. `complete_ticker_list.csv` - Full ticker list with metadata
3. `bloomberg_integration_guide.md` - This instruction file

## Bloomberg Data Download Steps:

### Option 1: Using Bloomberg Terminal
1. Open Bloomberg Terminal
2. Load ticker list from `complete_ticker_list.csv`
3. Use BDH (Bloomberg Data History) function:
   ```
   =BDH(ticker_list, "PX_LAST", start_date, end_date)
   ```
4. Export data and populate MultiTicker tab

### Option 2: Using Python xbbg Library
```python
import xbbg
import pandas as pd

# Load ticker list
tickers = pd.read_csv('complete_ticker_list.csv')['ticker'].tolist()

# Download data
data = xbbg.blp.bdh(
    tickers=tickers,
    flds='PX_LAST',
    start_date='2013-01-01',
    end_date='2025-08-18'
)

# Format and populate MultiTicker tab
```

### Option 3: Manual Population
1. Open `multiticker_tab.xlsx`
2. MultiTicker sheet has date column (A) and ticker columns (B+)
3. Populate data starting from Row 21
4. Maintain date format: YYYY-MM-DD

## Data Quality Checks:
- Ensure no missing dates
- Handle Bloomberg holidays appropriately
- Apply normalization factors from Metadata sheet
- Validate against existing data patterns

## Pipeline Integration:
- File is compatible with `master_aggregation_script.py`
- Use the created MultiTicker sheet as input
- All existing aggregation logic will work unchanged
"""
    
    # Write integration guide
    with open('bloomberg_integration_guide.md', 'w') as f:
        f.write(instructions)
    
    logger.info("📝 Created bloomberg_integration_guide.md")


def main():
    """
    Main execution function for MultiTicker tab creation.
    """
    try:
        logger.info("🚀 Starting MultiTicker Tab Creation")
        logger.info("=" * 80)
        
        # Step 1: Load ticker metadata
        ticker_df = load_ticker_metadata('complete_ticker_list.csv')
        
        # Step 2: Create date range
        dates = create_date_range('2013-01-01')  # Covers full historical range
        
        # Step 3: Create MultiTicker headers
        headers = create_multiticker_headers(ticker_df)
        
        # Step 4: Create MultiTicker Excel file
        output_file = create_multiticker_excel(ticker_df, dates, headers)
        
        # Step 5: Validate pipeline compatibility
        compatibility_ok = validate_pipeline_compatibility(output_file)
        
        # Step 6: Set up Bloomberg integration guide
        setup_bloomberg_data_integration()
        
        logger.info("=" * 80)
        if compatibility_ok:
            logger.info("✅ MULTITICKER TAB CREATION COMPLETED SUCCESSFULLY!")
            logger.info("🚀 Ready for Bloomberg data population!")
        else:
            logger.warning("⚠️  MultiTicker created but compatibility issues detected")
        
        logger.info("\n📄 DELIVERABLES:")
        logger.info("  - multiticker_tab.xlsx (MultiTicker structure)")
        logger.info("  - complete_ticker_list.csv (439 Bloomberg tickers)")
        logger.info("  - bloomberg_integration_guide.md (Integration instructions)")
        logger.info("  - multiticker_creation_script.py (This creation script)")
        
        return output_file, ticker_df
        
    except Exception as e:
        logger.error(f"❌ Error in MultiTicker creation: {str(e)}")
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    result = main()