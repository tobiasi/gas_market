#!/usr/bin/env python3
"""
INTEGRATED MASTER PIPELINE: European Gas Market Analysis
Complete end-to-end pipeline combining MultiTicker creation and gas demand processing.

This unified script:
1. Extracts Bloomberg ticker list from use4.xlsx
2. Creates MultiTicker tab structure with proper metadata
3. Provides Bloomberg data integration framework
4. Processes gas demand aggregation using existing logic
5. Generates validated European Gas Demand analysis

Combines functionality from:
- multiticker_creation_script.py (ticker extraction & tab creation)
- master_aggregation_script.py (gas demand processing)
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import logging
import warnings
import os
from typing import Tuple, Dict, List, Optional

warnings.filterwarnings('ignore')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# =============================================================================
# SECTION 1: TICKER EXTRACTION AND MULTITICKER CREATION
# =============================================================================

def extract_bloomberg_tickers_from_use4(file_path='use4.xlsx', sheet_name='TickerList') -> pd.DataFrame:
    """
    Extract complete Bloomberg ticker list from TickerList sheet with full metadata.
    """
    logger.info(f"Extracting Bloomberg tickers from {sheet_name} in {file_path}")
    
    # Load TickerList sheet - skip first 8 rows as indicated in CLAUDE.md
    df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=8)
    
    logger.info(f"Loaded TickerList with shape: {df.shape}")
    
    # Clean up column names
    df.columns = [str(col).strip() for col in df.columns]
    
    # Filter out rows with empty tickers
    original_count = len(df)
    df = df.dropna(subset=['Ticker'])
    df = df[df['Ticker'].astype(str).str.strip() != '']
    
    logger.info(f"Filtered from {original_count} to {len(df)} rows with valid tickers")
    
    # Validate Bloomberg tickers
    bloomberg_suffixes = ['Index', 'Comdty', 'BGN', 'Curncy', 'Govt', 'Corp', 'Equity']
    
    def is_bloomberg_ticker(ticker):
        if pd.isna(ticker):
            return False
        ticker_str = str(ticker).strip()
        return any(suffix in ticker_str for suffix in bloomberg_suffixes)
    
    df['Is_Bloomberg_Ticker'] = df['Ticker'].apply(is_bloomberg_ticker)
    bloomberg_df = df[df['Is_Bloomberg_Ticker']].copy()
    
    logger.info(f"Found {len(bloomberg_df)} Bloomberg tickers")
    
    # Process ticker metadata for MultiTicker format
    multiticker_data = []
    
    for idx, row in bloomberg_df.iterrows():
        ticker_info = {
            'ticker': str(row['Ticker']).strip(),
            'description': str(row.get('Description', '')).strip() if pd.notna(row.get('Description')) else '',
            'category': str(row.get('Category', '')).strip() if pd.notna(row.get('Category')) else '',
            'region_from': str(row.get('Region from', '')).strip() if pd.notna(row.get('Region from')) else '',
            'region_to': str(row.get('Region to', '')).strip() if pd.notna(row.get('Region to')) else '',
            'units': str(row.get('Units', 'GWh')).strip() if pd.notna(row.get('Units')) else 'GWh',
            'normalization_factor': row.get('Normalization factor', 1.0) if pd.notna(row.get('Normalization factor')) else 1.0,
            'positive_negative': str(row.get('Positive/Negative', '')).strip() if pd.notna(row.get('Positive/Negative')) else '',
            'start_date': str(row.get('Start date', '')).strip() if pd.notna(row.get('Start date')) else '',
            'other_notes': str(row.get('Other notes or comments', '')).strip() if pd.notna(row.get('Other notes or comments')) else ''
        }
        
        multiticker_data.append(ticker_info)
    
    ticker_df = pd.DataFrame(multiticker_data)
    
    # Round normalization factors
    ticker_df['normalization_factor'] = pd.to_numeric(ticker_df['normalization_factor'], errors='coerce').fillna(1.0).round(6)
    
    logger.info(f"Processed {len(ticker_df)} Bloomberg tickers for MultiTicker creation")
    
    return ticker_df


def create_multiticker_headers(ticker_df: pd.DataFrame) -> Dict[str, List[str]]:
    """
    Create the 3-row metadata headers for MultiTicker format.
    
    Row 14: Category (Demand, Import, Production, etc.)
    Row 15: Region (Country/area names)
    Row 16: Subcategory (Industrial, Gas-to-Power, LDZ, etc.)
    """
    logger.info("Creating MultiTicker 3-row headers")
    
    headers = {
        'categories': [],      # Row 14
        'regions': [],         # Row 15  
        'subcategories': []    # Row 16
    }
    
    for _, ticker_row in ticker_df.iterrows():
        # Row 14: Category (main type)
        category = ticker_row.get('category', '')
        headers['categories'].append(category)
        
        # Row 15: Region (geographic area)
        region = ticker_row.get('region_from', ticker_row.get('region_to', ''))
        headers['regions'].append(region)
        
        # Row 16: Subcategory (specific type) - derived from description and category
        description = str(ticker_row.get('description', '')).lower()
        category_lower = str(category).lower()
        
        subcategory = ''
        
        if 'industrial' in description:
            subcategory = 'Industrial'
        elif 'gas-to-power' in description or 'power' in description:
            subcategory = 'Gas-to-Power'
        elif 'ldz' in description or 'domestic' in description or 'residential' in description:
            subcategory = 'LDZ'
        elif 'storage' in description or 'inventory' in description:
            subcategory = 'Storage'
        elif 'lng' in description:
            subcategory = 'LNG'
        elif 'production' in category_lower:
            subcategory = 'Production'
        elif 'import' in category_lower:
            subcategory = 'Import'
        elif 'export' in category_lower:
            subcategory = 'Export'
        else:
            # Default based on category
            if 'demand' in category_lower:
                subcategory = 'Other'
            else:
                subcategory = category
        
        headers['subcategories'].append(subcategory)
    
    logger.info(f"Created headers for {len(ticker_df)} tickers")
    logger.info(f"  Categories: {len(set(headers['categories']))} unique")
    logger.info(f"  Regions: {len(set(headers['regions']))} unique")  
    logger.info(f"  Subcategories: {len(set(headers['subcategories']))} unique")
    
    return headers


def create_multiticker_structure(ticker_df: pd.DataFrame, 
                                headers: Dict[str, List[str]], 
                                start_date='2013-01-01',
                                end_date=None,
                                output_file='integrated_multiticker.xlsx') -> str:
    """
    Create the complete MultiTicker Excel structure with proper formatting.
    """
    logger.info(f"Creating MultiTicker structure: {output_file}")
    
    # Create date range
    if end_date is None:
        end_date = datetime.now().strftime('%Y-%m-%d')
    
    dates = pd.date_range(start=start_date, end=end_date, freq='D')
    logger.info(f"Date range: {start_date} to {end_date} ({len(dates)} dates)")
    
    # Create new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'MultiTicker'
    
    # Style definitions
    header_font = Font(bold=True, size=10)
    date_font = Font(size=9)
    
    # Row 14: Categories
    ws['A14'] = 'Date'
    ws['A14'].font = header_font
    
    for col_idx, category in enumerate(headers['categories'], start=2):
        ws.cell(row=14, column=col_idx, value=category).font = header_font
    
    # Row 15: Regions  
    for col_idx, region in enumerate(headers['regions'], start=2):
        ws.cell(row=15, column=col_idx, value=region).font = header_font
    
    # Row 16: Subcategories
    for col_idx, subcategory in enumerate(headers['subcategories'], start=2):
        ws.cell(row=16, column=col_idx, value=subcategory).font = header_font
    
    # Row 17: Ticker symbols (for reference)
    ws['A17'] = 'Ticker'
    ws['A17'].font = header_font
    
    for col_idx, ticker in enumerate(ticker_df['ticker'], start=2):
        ws.cell(row=17, column=col_idx, value=ticker).font = Font(size=8)
    
    # Row 21 onwards: Date data rows
    logger.info(f"Adding {len(dates)} date rows...")
    
    for date_idx, date in enumerate(dates, start=21):
        # Date column (Column A)
        ws.cell(row=date_idx, column=1, value=date).font = date_font
        
        # Data columns (B onwards) - initialized to 0 for data population
        for col_idx in range(2, len(ticker_df) + 2):
            ws.cell(row=date_idx, column=col_idx, value=0.0)
    
    # Format date column
    for row in range(21, 21 + len(dates)):
        ws.cell(row=row, column=1).number_format = 'YYYY-MM-DD'
    
    # Add metadata sheet for reference
    meta_ws = wb.create_sheet('Metadata')
    
    # Write ticker metadata to reference sheet
    for col_idx, col_name in enumerate(ticker_df.columns, start=1):
        meta_ws.cell(row=1, column=col_idx, value=col_name).font = header_font
    
    for row_idx, (_, row) in enumerate(ticker_df.iterrows(), start=2):
        for col_idx, value in enumerate(row, start=1):
            meta_ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Save workbook
    wb.save(output_file)
    logger.info(f"✅ MultiTicker structure created: {output_file}")
    
    return output_file


# =============================================================================
# SECTION 2: BLOOMBERG DATA INTEGRATION (PLACEHOLDER)
# =============================================================================

def populate_bloomberg_data(multiticker_file: str, 
                           ticker_df: pd.DataFrame, 
                           use_existing_data=True) -> str:
    """
    Populate MultiTicker with Bloomberg data.
    
    Options:
    1. use_existing_data=True: Use existing use4.xlsx MultiTicker sheet data
    2. use_existing_data=False: Placeholder for future Bloomberg API integration
    """
    logger.info("Populating MultiTicker with data...")
    
    if use_existing_data:
        logger.info("Using existing MultiTicker data from use4.xlsx")
        # For now, we'll use the existing populated use4.xlsx as the data source
        # This maintains compatibility with existing validation targets
        return 'use4.xlsx'
    else:
        logger.info("Bloomberg data integration placeholder")
        logger.warning("⚠️  Bloomberg API integration not implemented yet")
        logger.info("💡 For Bloomberg integration, see bloomberg_integration_guide.md")
        
        # Future Bloomberg integration would go here:
        # 1. Load ticker list
        # 2. Connect to Bloomberg API (xbbg library)
        # 3. Download historical data for all tickers
        # 4. Populate the MultiTicker structure
        # 5. Apply normalization factors
        # 6. Handle missing data
        
        return multiticker_file  # Return created structure for now


# =============================================================================
# SECTION 3: GAS DEMAND AGGREGATION PROCESSING
# =============================================================================

def load_multiticker_with_full_metadata(file_path='use4.xlsx', sheet_name='MultiTicker') -> Tuple[pd.DataFrame, Dict]:
    """
    Load MultiTicker data with complete metadata from rows 14, 15, 16.
    """
    logger.info(f"Loading MultiTicker with full metadata from {file_path}")
    
    # Load workbook to extract metadata
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    
    # Extract metadata from rows 14 (category), 15 (region), 16 (subcategory)
    metadata = {}
    max_col = min(ws.max_column, 600)  # Extended range
    
    logger.info(f"Extracting metadata from columns C to {openpyxl.utils.get_column_letter(max_col)}")
    
    for col in range(3, max_col + 1):  # Starting from column C
        col_name = f'Col_{col-2}'  # Col_1, Col_2, etc.
        category = ws.cell(row=14, column=col).value
        region = ws.cell(row=15, column=col).value  
        subcategory = ws.cell(row=16, column=col).value
        
        metadata[col_name] = {
            'category': str(category) if category else '',
            'region': str(region) if region else '',
            'subcategory': str(subcategory) if subcategory else ''
        }
    
    wb.close()
    
    # Load data using pandas
    df_full = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    
    # Data starts from row 21 (index 20), column B onwards
    data_rows = df_full.iloc[20:, 1:max_col].copy()
    
    # Set column names: Date + Col_1, Col_2, etc.
    data_rows.columns = ['Date'] + [f'Col_{i}' for i in range(1, len(data_rows.columns))]
    
    # Convert Date column
    data_rows['Date'] = pd.to_datetime(data_rows['Date'], errors='coerce')
    
    # Remove invalid dates
    data_rows = data_rows.dropna(subset=['Date'])
    
    # Convert data columns to numeric
    for col in data_rows.columns[1:]:
        data_rows[col] = pd.to_numeric(data_rows[col], errors='coerce')
    
    logger.info(f"Loaded {len(data_rows)} dates with {len(metadata)} tickers")
    
    return data_rows, metadata


def sumifs_three_criteria(data_df: pd.DataFrame, metadata: Dict, 
                         category_target: str, region_target: str, subcategory_target: str) -> pd.Series:
    """
    Exact replication of Excel 3-criteria SUMIFS with exact string matching.
    """
    matching_cols = []
    
    for col, info in metadata.items():
        if col in data_df.columns:
            # Exact string matching (case sensitive as in Excel)
            category_match = info['category'] == category_target
            region_match = info['region'] == region_target
            subcategory_match = info['subcategory'] == subcategory_target
            
            if category_match and region_match and subcategory_match:
                matching_cols.append(col)
    
    if not matching_cols:
        return pd.Series(0.0, index=data_df.index)
    
    # Sum across matching columns (handling NaN as 0)
    result = data_df[matching_cols].sum(axis=1, skipna=True)
    return result


def sumifs_two_criteria(data_df: pd.DataFrame, metadata: Dict, 
                       category_target: str, region_target: str) -> pd.Series:
    """
    Exact replication of Excel 2-criteria SUMIFS with exact string matching.
    """
    matching_cols = []
    
    for col, info in metadata.items():
        if col in data_df.columns:
            # Exact string matching (case sensitive as in Excel)
            category_match = info['category'] == category_target
            region_match = info['region'] == region_target
            
            if category_match and region_match:
                matching_cols.append(col)
    
    if not matching_cols:
        return pd.Series(0.0, index=data_df.index)
    
    # Sum across matching columns (handling NaN as 0)
    result = data_df[matching_cols].sum(axis=1, skipna=True)
    return result


def create_industrial_demand_sheet(data_df: pd.DataFrame, metadata: Dict) -> pd.DataFrame:
    """
    Task 4: Industrial Gas Demand calculation with exact Excel logic.
    Target: 240.70 for 2016-10-03
    """
    logger.info("Creating Industrial demand sheet (Task 4)")
    
    result = pd.DataFrame()
    result['Date'] = data_df['Date']
    
    # France Industrial
    result['France_Industrial'] = sumifs_three_criteria(
        data_df, metadata, 'Demand', 'France', 'Industrial'
    )
    
    # Belgium Industrial  
    result['Belgium_Industrial'] = sumifs_three_criteria(
        data_df, metadata, 'Demand', 'Belgium', 'Industrial'
    )
    
    # Italy Industrial
    result['Italy_Industrial'] = sumifs_three_criteria(
        data_df, metadata, 'Demand', 'Italy', 'Industrial'
    )
    
    # GB Industrial
    result['GB_Industrial'] = sumifs_three_criteria(
        data_df, metadata, 'Demand', 'GB', 'Industrial'
    )
    
    # Netherlands Industrial and Power
    result['Netherlands_IndPower'] = sumifs_three_criteria(
        data_df, metadata, 'Demand', 'Netherlands', 'Industrial and Power'
    )
    
    # Netherlands Zebra
    result['Netherlands_Zebra'] = sumifs_three_criteria(
        data_df, metadata, 'Demand', 'Netherlands', 'Zebra'
    )
    
    # Germany Industrial and Power (total)
    result['Germany_Total'] = sumifs_three_criteria(
        data_df, metadata, 'Demand', 'Germany', 'Industrial and Power'
    )
    
    # Germany Gas-to-Power (to subtract)
    result['Germany_GtP'] = sumifs_three_criteria(
        data_df, metadata, 'Intermediate Calculation', '#Germany', 'Gas-to-Power'
    )
    
    # Germany Industrial (calculated = Total - Gas-to-Power)
    result['Germany_Industrial'] = result['Germany_Total'] - result['Germany_GtP']
    
    # Total Industrial = Sum of all industrial components
    industrial_components = [
        'France_Industrial', 'Belgium_Industrial', 'Italy_Industrial', 'GB_Industrial',
        'Netherlands_IndPower', 'Netherlands_Zebra', 'Germany_Industrial'
    ]
    
    result['Total_Industrial_Demand'] = result[industrial_components].sum(axis=1)
    
    logger.info("Industrial demand sheet created successfully")
    return result


def create_ldz_demand_sheet(data_df: pd.DataFrame, metadata: Dict) -> pd.DataFrame:
    """
    Task 5: LDZ Demand calculation with exact Excel logic.
    Target: 307.80 for 2016-10-03
    """
    logger.info("Creating LDZ demand sheet (Task 5)")
    
    result = pd.DataFrame()
    result['Date'] = data_df['Date']
    
    ldz_total = pd.Series(0.0, index=data_df.index)
    
    # Standard LDZ countries
    ldz_countries = {
        'France': 'LDZ',
        'Belgium': 'LDZ', 
        'Italy': 'LDZ',
        'Netherlands': 'LDZ',
        'GB': 'LDZ',
        'Germany': 'LDZ'
    }
    
    for country, subcategory in ldz_countries.items():
        country_ldz = sumifs_three_criteria(
            data_df, metadata, 'Demand', country, subcategory
        )
        ldz_total += country_ldz
    
    # Italy Other
    italy_other = sumifs_three_criteria(
        data_df, metadata, 'Demand', 'Italy', 'Other'
    )
    ldz_total += italy_other
    
    # Special cases where subcategory matches region
    special_cases = {
        'Austria': 'Austria',
        'Switzerland': 'Switzerland', 
        'Luxembourg': 'Luxembourg'
    }
    
    for country, subcategory in special_cases.items():
        country_ldz = sumifs_three_criteria(
            data_df, metadata, 'Demand', country, subcategory
        )
        ldz_total += country_ldz
    
    # Ireland (using "Demand (Net)")
    ireland_ldz = sumifs_three_criteria(
        data_df, metadata, 'Demand (Net)', 'Island of Ireland', 'Island of Ireland'
    )
    ldz_total += ireland_ldz
    
    result['Total_LDZ_Demand'] = ldz_total
    
    logger.info("LDZ demand sheet created successfully")
    return result


def create_gas_to_power_demand_sheet(data_df: pd.DataFrame, metadata: Dict) -> pd.DataFrame:
    """
    Task 6: Gas-to-Power Demand calculation with BREAKTHROUGH solution.
    Key insight: Excel total excludes Netherlands from the sum.
    Target: 166.71 for 2016-10-03
    """
    logger.info("Creating Gas-to-Power demand sheet (Task 6 - Breakthrough)")
    
    result = pd.DataFrame()
    result['Date'] = data_df['Date']
    
    # Countries included in Excel total (excludes Netherlands)
    gtp_total = pd.Series(0.0, index=data_df.index)
    
    # Standard Gas-to-Power countries
    standard_gtp_countries = ['France', 'Belgium', 'Italy', 'GB']
    
    for country in standard_gtp_countries:
        country_gtp = sumifs_three_criteria(
            data_df, metadata, 'Demand', country, 'Gas-to-Power'
        )
        gtp_total += country_gtp
    
    # Germany Gas-to-Power (using Intermediate Calculation)
    germany_gtp = sumifs_three_criteria(
        data_df, metadata, 'Intermediate Calculation', '#Germany', 'Gas-to-Power'
    )
    gtp_total += germany_gtp
    
    # Netherlands is calculated but EXCLUDED from total (breakthrough insight)
    result['Total_Gas_to_Power_Demand'] = gtp_total
    
    logger.info("Gas-to-Power demand sheet created successfully")
    return result


def create_daily_country_demands(data_df: pd.DataFrame, metadata: Dict) -> pd.DataFrame:
    """
    Task 1: Country demand aggregation using 2-criteria SUMIFS.
    Target: France=90.13, Total=715.22 for 2016-10-03
    """
    logger.info("Creating daily country demands (Task 1)")
    
    result = pd.DataFrame()
    result['Date'] = data_df['Date']
    
    # Define countries to process
    countries = [
        'France', 'Belgium', 'Italy', 'Netherlands', 'GB', 
        'Austria', 'Germany', 'Switzerland', 'Luxembourg'
    ]
    
    country_totals = []
    
    for country in countries:
        country_demand = sumifs_two_criteria(
            data_df, metadata, 'Demand', country
        )
        result[country] = country_demand
        country_totals.append(country_demand)
    
    # Ireland using "Demand (Net)"
    ireland_demand = sumifs_two_criteria(
        data_df, metadata, 'Demand (Net)', 'Island of Ireland'
    )
    result['Ireland'] = ireland_demand
    country_totals.append(ireland_demand)
    
    # Calculate total
    total_demand = pd.Series(0.0, index=data_df.index)
    for country_total in country_totals:
        total_demand += country_total
    
    result['Total'] = total_demand
    
    logger.info("Daily country demands created successfully")
    return result


def merge_all_components(daily_country_data: pd.DataFrame, 
                        industrial_data: pd.DataFrame, 
                        ldz_data: pd.DataFrame, 
                        gtp_data: pd.DataFrame) -> pd.DataFrame:
    """
    Merge all components into complete Daily historic data.
    """
    logger.info("Merging all components into complete dataset")
    
    # Start with country data
    complete_data = daily_country_data.copy()
    
    # Add subcategory totals
    complete_data['Industrial'] = industrial_data['Total_Industrial_Demand']
    complete_data['LDZ'] = ldz_data['Total_LDZ_Demand'] 
    complete_data['Gas_to_Power'] = gtp_data['Total_Gas_to_Power_Demand']
    
    # Sort by date
    complete_data = complete_data.sort_values('Date').reset_index(drop=True)
    
    logger.info("Component merging completed successfully")
    return complete_data


def validate_complete_results(complete_data: pd.DataFrame) -> bool:
    """
    Validate against known Excel values for comprehensive testing.
    """
    logger.info("Running comprehensive validation suite")
    
    # Target values for 2016-10-03
    targets = {
        'France': 90.13,
        'Total': 715.22,
        'Industrial': 240.70,
        'LDZ': 307.80,
        'Gas_to_Power': 166.71
    }
    
    # Find validation date
    validation_date = '2016-10-03'
    sample = complete_data[complete_data['Date'] == validation_date]
    
    if sample.empty:
        logger.warning(f"Validation date {validation_date} not found in results")
        return False
    
    logger.info(f"\n📊 INTEGRATED PIPELINE VALIDATION for {validation_date}:")
    logger.info("=" * 70)
    
    all_pass = True
    sample_row = sample.iloc[0]
    
    for column, target in targets.items():
        if column in sample.columns:
            actual = sample_row[column]
            diff = abs(actual - target)
            
            if diff < 0.01:
                status = "✅ PERFECT"
            elif diff < 0.5:
                status = "✅ GOOD"  
            else:
                status = "❌ FAIL"
                all_pass = False
            
            logger.info(f"  {status} {column}: {actual:.2f} (target {target:.2f}, diff: {diff:.2f})")
        else:
            logger.warning(f"  ❌ {column}: Column not found")
            all_pass = False
    
    logger.info("=" * 70)
    
    if all_pass:
        logger.info("🎯 INTEGRATED PIPELINE VALIDATION SUCCESS!")
        logger.info("🚀 All components working perfectly together!")
    else:
        logger.warning("⚠️  Some validation targets not met")
    
    return all_pass


# =============================================================================
# SECTION 4: INTEGRATED MASTER PIPELINE ORCHESTRATION
# =============================================================================

def process_gas_demand_aggregation(data_source_file: str = 'use4.xlsx') -> pd.DataFrame:
    """
    Process gas demand aggregation using existing validation logic.
    """
    logger.info("🔄 Starting gas demand aggregation processing...")
    
    # Load MultiTicker with full metadata
    data_df, metadata = load_multiticker_with_full_metadata(data_source_file, 'MultiTicker')
    
    # Create all subcategory sheets
    logger.info("🏭 Creating Industrial demand sheet...")
    industrial_data = create_industrial_demand_sheet(data_df, metadata)
    
    logger.info("🏠 Creating LDZ demand sheet...")
    ldz_data = create_ldz_demand_sheet(data_df, metadata)
    
    logger.info("⚡ Creating Gas-to-Power demand sheet...")
    gtp_data = create_gas_to_power_demand_sheet(data_df, metadata)
    
    # Create country demand data
    logger.info("🌍 Creating country demand aggregation...")
    daily_country_data = create_daily_country_demands(data_df, metadata)
    
    # Merge everything
    logger.info("🔗 Merging all components...")
    complete_data = merge_all_components(daily_country_data, industrial_data, ldz_data, gtp_data)
    
    return complete_data


def main(create_structure: bool = True,
         populate_data: bool = False,
         process_aggregation: bool = True,
         use_existing_data: bool = True,
         output_prefix: str = 'integrated') -> Optional[pd.DataFrame]:
    """
    Integrated Master Pipeline - Complete European Gas Market Analysis
    
    Parameters:
    - create_structure: Create MultiTicker tab structure from ticker list
    - populate_data: Populate with Bloomberg data (placeholder)
    - process_aggregation: Run gas demand aggregation analysis
    - use_existing_data: Use existing use4.xlsx data for processing
    - output_prefix: Prefix for output files
    """
    
    logger.info("🚀 STARTING INTEGRATED MASTER PIPELINE")
    logger.info("=" * 90)
    logger.info("European Gas Market Analysis - Complete End-to-End Processing")
    logger.info("=" * 90)
    
    try:
        data_source_file = 'use4.xlsx'  # Default data source
        
        # =================================================================
        # PHASE 1: MultiTicker Structure Creation
        # =================================================================
        if create_structure:
            logger.info("📋 PHASE 1: Creating MultiTicker Structure")
            logger.info("-" * 50)
            
            # Step 1: Extract Bloomberg tickers
            logger.info("📊 Step 1.1: Extracting Bloomberg tickers...")
            ticker_df = extract_bloomberg_tickers_from_use4('use4.xlsx', 'TickerList')
            
            # Export ticker list for reference
            ticker_csv = f'{output_prefix}_ticker_list.csv'
            ticker_df.to_csv(ticker_csv, index=False)
            logger.info(f"💾 Exported ticker list: {ticker_csv}")
            
            # Step 2: Create MultiTicker headers
            logger.info("🏗️  Step 1.2: Creating MultiTicker headers...")
            headers = create_multiticker_headers(ticker_df)
            
            # Step 3: Create MultiTicker structure
            logger.info("📋 Step 1.3: Creating MultiTicker Excel structure...")
            multiticker_file = create_multiticker_structure(
                ticker_df, headers, output_file=f'{output_prefix}_multiticker.xlsx'
            )
            
            logger.info("✅ Phase 1 completed: MultiTicker structure created")
        
        # =================================================================
        # PHASE 2: Bloomberg Data Population (Optional)
        # =================================================================
        if populate_data:
            logger.info("\n📡 PHASE 2: Bloomberg Data Population")
            logger.info("-" * 50)
            
            data_source_file = populate_bloomberg_data(
                multiticker_file if create_structure else f'{output_prefix}_multiticker.xlsx',
                ticker_df if create_structure else pd.DataFrame(),
                use_existing_data=use_existing_data
            )
            
            logger.info("✅ Phase 2 completed: Data population ready")
        
        # =================================================================
        # PHASE 3: Gas Demand Aggregation Processing
        # =================================================================
        if process_aggregation:
            logger.info("\n⚙️  PHASE 3: Gas Demand Aggregation Processing")
            logger.info("-" * 50)
            
            # Process the complete gas demand aggregation
            complete_data = process_gas_demand_aggregation(data_source_file)
            
            # Validate results
            logger.info("🔍 Running validation...")
            validation_passed = validate_complete_results(complete_data)
            
            if validation_passed:
                # Export final results
                logger.info("📊 Exporting final results...")
                
                # Format dates for export
                export_data = complete_data.copy()
                export_data['Date'] = export_data['Date'].dt.strftime('%Y-%m-%d')
                
                # Round to match Excel precision
                numeric_cols = export_data.select_dtypes(include=[np.number]).columns
                export_data[numeric_cols] = export_data[numeric_cols].round(2)
                
                # Export to CSV
                output_file = f'{output_prefix}_daily_historic_data_by_category.csv'
                export_data.to_csv(output_file, index=False)
                
                logger.info(f"✅ Final output exported: {output_file}")
                logger.info("✅ Phase 3 completed: Gas demand aggregation processed")
                
                # Show sample results
                logger.info(f"\n📋 Sample results for 2016-10-03:")
                sample = export_data[export_data['Date'] == '2016-10-03']
                if not sample.empty:
                    sample_row = sample.iloc[0]
                    logger.info(f"  🇫🇷 France: {sample_row['France']}")
                    logger.info(f"  📊 Total: {sample_row['Total']}")
                    logger.info(f"  🏭 Industrial: {sample_row['Industrial']}")
                    logger.info(f"  🏠 LDZ: {sample_row['LDZ']}")
                    logger.info(f"  ⚡ Gas-to-Power: {sample_row['Gas_to_Power']}")
                
                return complete_data
            else:
                logger.error("❌ Validation failed - pipeline aborted")
                return None
        
        # =================================================================
        # COMPLETION
        # =================================================================
        logger.info("\n" + "=" * 90)
        logger.info("🎯 INTEGRATED MASTER PIPELINE COMPLETED SUCCESSFULLY!")
        logger.info("🚀 Complete European Gas Market Analysis Ready!")
        
        logger.info(f"\n📄 DELIVERABLES CREATED:")
        if create_structure:
            logger.info(f"  - {output_prefix}_multiticker.xlsx (MultiTicker structure)")
            logger.info(f"  - {output_prefix}_ticker_list.csv (Bloomberg ticker list)")
        if process_aggregation:
            logger.info(f"  - {output_prefix}_daily_historic_data_by_category.csv (Final analysis)")
        
        logger.info("\n💡 INTEGRATION OPTIONS:")
        logger.info("  - Bloomberg Terminal: Use BDH function with ticker list")
        logger.info("  - Python xbbg: Automated Bloomberg API integration") 
        logger.info("  - Manual: Populate MultiTicker structure directly")
        
        logger.info("=" * 90)
        
        return complete_data if process_aggregation else None
        
    except Exception as e:
        logger.error(f"❌ Integrated pipeline failed: {str(e)}")
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    # Run the complete integrated pipeline
    result = main(
        create_structure=True,      # Create MultiTicker infrastructure
        populate_data=False,        # Skip Bloomberg population (placeholder)
        process_aggregation=True,   # Run full gas demand analysis
        use_existing_data=True,     # Use existing use4.xlsx data
        output_prefix='integrated'  # Prefix for output files
    )