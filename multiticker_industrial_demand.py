#!/usr/bin/env python3
"""
Create Industrial gas demand sheet with 3-criteria SUMIFS logic.
Also handles LDZ and Gas-to-Power subcategory aggregations.
"""

import pandas as pd
import numpy as np
from datetime import datetime
import logging
import openpyxl
import warnings
warnings.filterwarnings('ignore')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def load_multiticker_with_subcategories(file_path='use4.xlsx', sheet_name='MultiTicker'):
    """
    Load MultiTicker data with 3-criteria metadata (including row 16 subcategories).
    
    Returns:
        data_df: DataFrame with dates and ticker values
        metadata: Dictionary with category, region, and subcategory for each ticker
    """
    logger.info(f"Loading MultiTicker sheet with subcategories from {file_path}")
    
    # Load workbook to get metadata
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    
    # Extract metadata from rows 14, 15, 16
    metadata = {}
    max_col = min(ws.max_column, 500)  # Limit for performance
    
    for col in range(3, max_col + 1):  # Starting from column C
        col_name = f'Col_{col-3}'
        category = ws.cell(row=14, column=col).value
        region = ws.cell(row=15, column=col).value
        subcategory = ws.cell(row=16, column=col).value
        
        metadata[col_name] = {
            'category': str(category) if category else '',
            'region': str(region) if region else '',
            'subcategory': str(subcategory) if subcategory else ''
        }
    
    wb.close()
    
    # Now load the data using pandas
    df_full = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    
    # Data starts from row 21 (index 20)
    data_rows = df_full.iloc[20:, 1:max_col].copy()  # Column B onwards
    
    # Set column names
    data_rows.columns = ['Date'] + [f'Col_{i}' for i in range(len(data_rows.columns)-1)]
    
    # Convert first column to datetime
    data_rows['Date'] = pd.to_datetime(data_rows['Date'], errors='coerce')
    
    # Remove rows with invalid dates
    data_rows = data_rows.dropna(subset=['Date'])
    
    # Convert data columns to numeric
    for col in data_rows.columns[1:]:
        data_rows[col] = pd.to_numeric(data_rows[col], errors='coerce')
    
    logger.info(f"Loaded {len(data_rows)} dates with {len(metadata)} tickers")
    
    # Log subcategory summary
    subcats = set(m['subcategory'] for m in metadata.values() if m['subcategory'])
    logger.info(f"Found {len(subcats)} unique subcategories")
    
    return data_rows, metadata


def aggregate_with_three_criteria(data_df, metadata, target_category, target_region, target_subcategory):
    """
    Aggregate data using 3-criteria SUMIFS logic.
    
    Args:
        target_category: Category to match (e.g., 'Demand')
        target_region: Region to match (e.g., 'France')
        target_subcategory: Subcategory to match (e.g., 'Industrial')
    
    Returns:
        Series with aggregated values for each date
    """
    matching_cols = []
    
    for col, info in metadata.items():
        # Check all three criteria
        cat_match = info['category'] == target_category
        reg_match = info['region'] == target_region
        
        # Handle subcategory matching (case-insensitive and flexible)
        subcat = info['subcategory'].lower()
        target_sub = target_subcategory.lower()
        
        # Special handling for composite subcategories
        if target_sub == 'industrial':
            subcat_match = 'industrial' in subcat
        elif target_sub == 'gas-to-power':
            subcat_match = 'gas-to-power' in subcat or 'gas to power' in subcat
        elif target_sub == 'ldz':
            subcat_match = 'ldz' in subcat
        else:
            subcat_match = target_sub in subcat
        
        if cat_match and reg_match and subcat_match:
            matching_cols.append(col)
    
    if not matching_cols:
        return pd.Series(0, index=data_df.index)
    
    logger.debug(f"Found {len(matching_cols)} columns for {target_category}/{target_region}/{target_subcategory}")
    
    # Sum across matching columns
    result = data_df[matching_cols].sum(axis=1, skipna=True)
    
    return result


def calculate_industrial_demand(data_df, metadata):
    """
    Calculate industrial demand for all countries following Excel logic.
    
    Returns DataFrame with columns:
    - Date
    - France_Industrial, Belgium_Industrial, Italy_Industrial, GB_Industrial
    - Netherlands_Industrial, Netherlands_Zebra
    - Germany_Total, Germany_GtP, Germany_Industrial
    - Total_Industrial
    """
    logger.info("Calculating industrial demand by country")
    
    result = pd.DataFrame()
    result['Date'] = data_df['Date']
    
    # Standard countries with Industrial subcategory
    standard_countries = ['France', 'Belgium', 'Italy', 'GB']
    
    for country in standard_countries:
        result[f'{country}_Industrial'] = aggregate_with_three_criteria(
            data_df, metadata, 'Demand', country, 'Industrial'
        )
    
    # Netherlands special handling
    # Component 1: Industrial and Power
    result['Netherlands_IndPower'] = aggregate_with_three_criteria(
        data_df, metadata, 'Demand', 'Netherlands', 'Industrial and Power'
    )
    
    # Component 2: Zebra
    result['Netherlands_Zebra'] = aggregate_with_three_criteria(
        data_df, metadata, 'Demand', 'Netherlands', 'Zebra'
    )
    
    # Germany special calculation
    # Total demand (Industrial and Power)
    result['Germany_Total'] = aggregate_with_three_criteria(
        data_df, metadata, 'Demand', 'Germany', 'Industrial and Power'
    )
    
    # Gas-to-Power (from Intermediate Calculation)
    result['Germany_GtP'] = aggregate_with_three_criteria(
        data_df, metadata, 'Intermediate Calculation', '#Germany', 'Gas-to-Power'
    )
    
    # Germany Industrial = Total - Gas-to-Power
    result['Germany_Industrial'] = result['Germany_Total'] - result['Germany_GtP']
    
    # Calculate total industrial demand
    industrial_cols = [
        'France_Industrial', 'Belgium_Industrial', 'Italy_Industrial', 'GB_Industrial',
        'Netherlands_IndPower', 'Netherlands_Zebra', 'Germany_Industrial'
    ]
    
    result['Total_Industrial'] = result[industrial_cols].sum(axis=1)
    
    # Sort by date
    result = result.sort_values('Date').reset_index(drop=True)
    
    logger.info(f"Calculated industrial demand for {len(result)} dates")
    
    return result


def calculate_ldz_demand(data_df, metadata):
    """
    Calculate LDZ (Local Distribution Zone) demand across all countries.
    """
    logger.info("Calculating LDZ demand")
    
    result = pd.DataFrame()
    result['Date'] = data_df['Date']
    
    # Find all LDZ-related demand
    ldz_total = pd.Series(0, index=data_df.index, dtype=float)
    
    # Check each country for LDZ demand
    countries = ['France', 'Belgium', 'Italy', 'Netherlands', 'GB', 
                'Austria', 'Germany', 'Switzerland', 'Luxembourg']
    
    for country in countries:
        country_ldz = aggregate_with_three_criteria(
            data_df, metadata, 'Demand', country, 'LDZ'
        )
        if country_ldz.sum() > 0:
            result[f'{country}_LDZ'] = country_ldz
            ldz_total += country_ldz
            logger.info(f"Found LDZ data for {country}")
    
    result['Total_LDZ'] = ldz_total
    
    return result


def calculate_gas_to_power_demand(data_df, metadata):
    """
    Calculate Gas-to-Power demand across all countries.
    """
    logger.info("Calculating Gas-to-Power demand")
    
    result = pd.DataFrame()
    result['Date'] = data_df['Date']
    
    # Find all Gas-to-Power demand
    gtp_total = pd.Series(0, index=data_df.index, dtype=float)
    
    # Check each country for Gas-to-Power demand
    countries = ['France', 'Belgium', 'Italy', 'Netherlands', 'GB', 
                'Austria', 'Germany', 'Switzerland', 'Luxembourg']
    
    for country in countries:
        country_gtp = aggregate_with_three_criteria(
            data_df, metadata, 'Demand', country, 'Gas-to-Power'
        )
        if country_gtp.sum() > 0:
            result[f'{country}_GtP'] = country_gtp
            gtp_total += country_gtp
            logger.info(f"Found Gas-to-Power data for {country}")
    
    # Also add the Germany Gas-to-Power from Intermediate Calculation
    germany_gtp_special = aggregate_with_three_criteria(
        data_df, metadata, 'Intermediate Calculation', '#Germany', 'Gas-to-Power'
    )
    if germany_gtp_special.sum() > 0:
        gtp_total += germany_gtp_special
        result['Germany_GtP_Special'] = germany_gtp_special
    
    result['Total_Gas_to_Power'] = gtp_total
    
    return result


def create_subcategory_totals(data_df, metadata):
    """
    Create all subcategory totals (Industrial, LDZ, Gas-to-Power).
    
    Returns DataFrame with Date and three total columns.
    """
    logger.info("Creating subcategory totals")
    
    # Calculate each subcategory
    industrial = calculate_industrial_demand(data_df, metadata)
    ldz = calculate_ldz_demand(data_df, metadata)
    gtp = calculate_gas_to_power_demand(data_df, metadata)
    
    # Combine into single DataFrame
    result = pd.DataFrame()
    result['Date'] = data_df['Date']
    result['Industrial'] = industrial['Total_Industrial']
    result['LDZ'] = ldz['Total_LDZ']
    result['Gas_to_Power'] = gtp['Total_Gas_to_Power']
    
    # Sort by date
    result = result.sort_values('Date').reset_index(drop=True)
    
    return result, industrial, ldz, gtp


def validate_industrial_values(industrial_df):
    """
    Validate industrial calculations against known Excel values.
    """
    logger.info("Validating industrial calculations")
    
    # Known values from Excel
    validations = [
        {'date': '2016-10-03', 'Total_Industrial': 228.80},
        {'date': '2016-10-04', 'Total_Industrial': 268.26}
    ]
    
    for val in validations:
        date = pd.to_datetime(val['date'])
        row = industrial_df[industrial_df['Date'] == date]
        
        if row.empty:
            logger.warning(f"Date {val['date']} not found")
            continue
        
        actual = row['Total_Industrial'].iloc[0]
        expected = val['Total_Industrial']
        diff = abs(actual - expected)
        
        if diff < 0.5:  # Allow small tolerance
            logger.info(f"✓ {val['date']}: Industrial = {actual:.2f} (expected {expected:.2f})")
        else:
            logger.warning(f"✗ {val['date']}: Industrial = {actual:.2f} (expected {expected:.2f}, diff: {diff:.2f})")


def export_subcategory_data(subcategory_df, output_file='subcategory_totals.csv'):
    """
    Export subcategory totals to CSV.
    """
    # Format date column
    subcategory_df = subcategory_df.copy()
    subcategory_df['Date'] = subcategory_df['Date'].dt.strftime('%Y-%m-%d')
    
    # Round numeric columns
    numeric_cols = subcategory_df.select_dtypes(include=[np.number]).columns
    subcategory_df[numeric_cols] = subcategory_df[numeric_cols].round(2)
    
    # Export
    subcategory_df.to_csv(output_file, index=False)
    logger.info(f"Exported subcategory totals to {output_file}")
    
    return output_file


def main():
    """
    Main execution function.
    """
    try:
        # Load MultiTicker data with subcategories
        data_df, metadata = load_multiticker_with_subcategories('use4.xlsx', 'MultiTicker')
        
        # Create all subcategory totals
        subcategory_totals, industrial, ldz, gtp = create_subcategory_totals(data_df, metadata)
        
        # Validate industrial calculations
        validate_industrial_values(industrial)
        
        # Display sample results
        logger.info("\nSubcategory totals for 2016-10-03:")
        sample = subcategory_totals[subcategory_totals['Date'] == '2016-10-03']
        if not sample.empty:
            print(f"  Industrial: {sample['Industrial'].iloc[0]:.2f}")
            print(f"  LDZ: {sample['LDZ'].iloc[0]:.2f}")
            print(f"  Gas-to-Power: {sample['Gas_to_Power'].iloc[0]:.2f}")
        
        # Export results
        export_subcategory_data(subcategory_totals)
        
        # Also export detailed industrial breakdown
        industrial[['Date', 'France_Industrial', 'Belgium_Industrial', 'Italy_Industrial', 
                  'GB_Industrial', 'Germany_Industrial', 'Total_Industrial']].to_csv(
            'industrial_breakdown.csv', index=False
        )
        logger.info("Exported industrial breakdown to industrial_breakdown.csv")
        
        return subcategory_totals, industrial
        
    except Exception as e:
        logger.error(f"Error in main execution: {str(e)}")
        raise


if __name__ == "__main__":
    subcategory_totals, industrial_details = main()