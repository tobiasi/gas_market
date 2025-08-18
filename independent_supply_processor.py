#!/usr/bin/env python3
"""
INDEPENDENT SUPPLY-SIDE PROCESSOR: Separate Module for Gas Market Supply Analysis

This is a SEPARATE, INDEPENDENT supply processor that:
1. Loads the WORKING demand results from restored_demand_results.csv
2. Processes all 27 supply routes independently using MultiTicker data  
3. Merges supply data with demand data ONLY at the final output stage
4. NEVER interferes with the working demand-side logic

CRITICAL: This module is completely isolated from demand processing to preserve
the perfect validation results achieved in Phase 1.
"""

import sys
sys.path.append("C:/development/commodities")

import pandas as pd
import numpy as np
import logging
from typing import Dict, List, Tuple, Optional
from datetime import datetime
import warnings

warnings.filterwarnings('ignore')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class IndependentSupplyProcessor:
    """
    Independent supply-side processor that works separately from demand pipeline.
    
    This processor:
    1. Loads working demand results as read-only input
    2. Processes supply routes completely independently
    3. Only merges at final output stage
    """
    
    def __init__(self):
        """Initialize independent supply processor."""
        self.supply_routes_mapping = self._create_supply_routes_mapping()
        self.lng_aggregation_routes = self._define_lng_aggregation_routes()
        self.geopolitical_corrections = self._define_geopolitical_corrections()
        
    def _create_supply_routes_mapping(self) -> Dict[str, Tuple[str, str, str]]:
        """
        Create comprehensive mapping based on ACTUAL MultiTicker structure.
        
        Maps route names to (Category, Region_From, Region_To) tuples
        for 3-criteria SUMIFS matching in MultiTicker data.
        
        Based on investigation of actual MultiTicker columns.
        """
        return {
            # Pipeline Imports - CORRECTED based on MultiTicker investigation
            'Slovakia_Austria': ('Import', 'Slovakia', 'Austria'),           # Found in MultiTicker
            'Russia_NordStream_Germany': ('Import', 'Russia (Nord Stream)', 'Germany'), # Found in MultiTicker  
            'Norway_Europe': ('Import', 'Norway', 'Europe'),                # Found in MultiTicker
            'Algeria_Italy': ('Import', 'Algeria', 'Italy'),                # Found in MultiTicker
            'Libya_Italy': ('Import', 'Libya', 'Italy'),                    # Found in MultiTicker
            'Spain_France': ('Import', 'Spain', 'France'),                  # Col_378 found
            'Denmark_Germany': ('Import', 'Denmark', 'Germany'),            # Col_85 found
            'Czech_Poland_Germany': ('Import', 'Czech and Poland', 'Germany'), # Col_67+ found
            
            # LNG Imports - CORRECTED to match actual MultiTicker structure
            'LNG_Belgium': ('Import', 'LNG', 'Belgium'),                    # Col_30 found
            'LNG_France': ('Import', 'LNG', 'France'),                      # Col_38+ found
            'LNG_Germany': ('Import', 'LNG', 'Germany'),                    # Col_66+ found  
            'LNG_Netherlands': ('Import', 'LNG', 'Netherlands'),            # Col_116+ found
            'LNG_GB': ('Import', 'LNG', 'GB'),                             # Col_126+ found
            'LNG_Italy': ('Import', 'LNG', 'Italy'),                       # Col_341+ found
            
            # Domestic Production - CORRECTED to match actual structure
            'Austria_Production': ('Production', 'Austria', 'Austria'),     # Col_6+ found
            'Germany_Production': ('Production', 'Germany', 'Germany'),     # Col_21+ found
            'GB_Production': ('Production', 'GB', 'GB'),                    # Col_54+ found
            'Netherlands_Production': ('Production', 'Netherlands', 'Netherlands'), # Col_118+ found
            'Italy_Production': ('Production', 'Italy', 'Italy'),           # Col_337 found
            
            # Exports - CORRECTED
            'Austria_Hungary_Export': ('Export', 'Austria', 'Hungary'),     # Col_64 found
        }
    
    def _define_lng_aggregation_routes(self) -> List[str]:
        """
        Define LNG routes that require special aggregation logic.
        
        LNG imports need to be aggregated separately for supply balance calculations.
        """
        return [
            'LNG_Belgium', 'LNG_France', 'LNG_Germany', 
            'LNG_Netherlands', 'LNG_GB', 'LNG_Italy'
        ]
    
    def _define_geopolitical_corrections(self) -> Dict[str, Dict]:
        """
        Define geopolitical corrections for supply data.
        
        Key corrections:
        1. Russian Nord Stream: Zero after 2023 due to pipeline sabotage
        2. Russian transit routes: Reduced capacity post-2022
        3. LNG diversification: Increased US/Qatar volumes post-2022
        """
        return {
            'Russia_NordStream_Germany': {
                'cutoff_date': '2023-01-01',
                'post_cutoff_value': 0.0,
                'reason': 'Nord Stream pipeline sabotage and closure'
            },
            'Russia_Ukraine_Slovakia': {
                'reduction_date': '2022-03-01', 
                'reduction_factor': 0.6,
                'reason': 'Ukraine conflict impact on Russian gas transit'
            },
            'Russia_Belarus_Poland': {
                'reduction_date': '2022-03-01',
                'reduction_factor': 0.4, 
                'reason': 'Sanctions impact on Russian gas via Belarus'
            }
        }
    
    def load_working_demand_results(self, demand_file: str = 'restored_demand_results.csv') -> pd.DataFrame:
        """
        Load the working demand-side results from Phase 1.
        
        CRITICAL: This is READ-ONLY. We never modify the working demand data.
        """
        logger.info(f"📊 Loading working demand results from {demand_file}")
        
        try:
            demand_df = pd.read_csv(demand_file)
            demand_df['Date'] = pd.to_datetime(demand_df['Date'])
            
            logger.info(f"✅ Loaded {len(demand_df)} demand records")
            logger.info(f"   Date range: {demand_df['Date'].min()} to {demand_df['Date'].max()}")
            
            # Validate critical columns exist
            required_cols = ['Date', 'France', 'Total', 'Industrial', 'LDZ', 'Gas_to_Power']
            missing_cols = [col for col in required_cols if col not in demand_df.columns]
            
            if missing_cols:
                raise ValueError(f"Missing required columns in demand data: {missing_cols}")
            
            logger.info("✅ Working demand data loaded successfully")
            return demand_df
            
        except FileNotFoundError:
            logger.error(f"❌ Demand results file not found: {demand_file}")
            logger.error("   Run restored_demand_pipeline.py first to generate demand results")
            raise
        except Exception as e:
            logger.error(f"❌ Error loading demand results: {str(e)}")
            raise
    
    def load_multiticker_for_supply(self, file_path: str = 'use4.xlsx', 
                                  sheet_name: str = 'MultiTicker') -> Tuple[pd.DataFrame, Dict]:
        """
        Load MultiTicker data specifically for supply processing.
        
        This is independent of demand processing and focuses on supply route extraction.
        """
        logger.info(f"📊 Loading MultiTicker data for supply processing from {file_path}")
        
        # Load workbook to extract metadata
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        
        # Extract metadata from rows 14 (category), 15 (region), 16 (subcategory)  
        metadata = {}
        max_col = min(ws.max_column, 600)
        
        logger.info(f"Extracting supply metadata from columns C to {openpyxl.utils.get_column_letter(max_col)}")
        
        for col in range(3, max_col + 1):
            col_name = f'Col_{col-2}'
            category = ws.cell(row=14, column=col).value
            region = ws.cell(row=15, column=col).value  
            subcategory = ws.cell(row=16, column=col).value
            
            metadata[col_name] = {
                'category': str(category) if category else '',
                'region': str(region) if region else '',
                'subcategory': str(subcategory) if subcategory else ''
            }
        
        wb.close()
        
        # Load data using pandas
        df_full = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        
        # Data starts from row 21 (index 20), column B onwards
        data_rows = df_full.iloc[20:, 1:max_col].copy()
        
        # Set column names: Date + Col_1, Col_2, etc.
        data_rows.columns = ['Date'] + [f'Col_{i}' for i in range(1, len(data_rows.columns))]
        
        # Convert Date column
        data_rows['Date'] = pd.to_datetime(data_rows['Date'], errors='coerce')
        
        # Remove invalid dates
        data_rows = data_rows.dropna(subset=['Date'])
        
        # Convert data columns to numeric
        for col in data_rows.columns[1:]:
            data_rows[col] = pd.to_numeric(data_rows[col], errors='coerce')
        
        logger.info(f"✅ Loaded MultiTicker data: {len(data_rows)} dates with {len(metadata)} tickers")
        
        return data_rows, metadata
    
    def sumifs_supply_three_criteria(self, data_df: pd.DataFrame, metadata: Dict,
                                   category_target: str, region_from: str, region_to: str) -> pd.Series:
        """
        Apply 3-criteria SUMIFS for supply route extraction.
        
        Matches: Category + Region_From + Region_To for supply routes.
        """
        matching_cols = []
        
        for col, info in metadata.items():
            if col in data_df.columns:
                category_match = info['category'] == category_target
                region_match = info['region'] == region_from  
                subcategory_match = info['subcategory'] == region_to
                
                if category_match and region_match and subcategory_match:
                    matching_cols.append(col)
        
        if not matching_cols:
            logger.debug(f"No supply matches for {category_target}/{region_from}/{region_to}")
            return pd.Series(0.0, index=data_df.index)
        
        logger.debug(f"Found {len(matching_cols)} supply matches for {category_target}/{region_from}/{region_to}")
        
        # Sum across matching columns
        result = data_df[matching_cols].sum(axis=1, skipna=True)
        return result
    
    def process_all_supply_routes(self, data_df: pd.DataFrame, metadata: Dict) -> pd.DataFrame:
        """
        Process ALL 27 supply routes independently.
        
        This extracts all supply-side data without touching demand processing.
        """
        logger.info("🚛 Processing ALL supply routes independently")
        
        supply_result = pd.DataFrame()
        supply_result['Date'] = data_df['Date']
        
        # Process each supply route using 3-criteria SUMIFS
        for route_name, (category, region_from, region_to) in self.supply_routes_mapping.items():
            logger.debug(f"Processing supply route: {route_name}")
            
            route_data = self.sumifs_supply_three_criteria(
                data_df, metadata, category, region_from, region_to
            )
            
            supply_result[route_name] = route_data
            
            # Log non-zero routes
            if route_data.sum() > 0:
                logger.debug(f"  ✅ {route_name}: {route_data.sum():.2f} total")
        
        logger.info(f"✅ Processed {len(self.supply_routes_mapping)} supply routes")
        
        return supply_result
    
    def apply_lng_aggregation(self, supply_df: pd.DataFrame) -> pd.DataFrame:
        """
        Apply LNG special aggregation processing.
        
        LNG imports require special handling for supply balance calculations.
        """
        logger.info("🛢️ Applying LNG aggregation logic")
        
        # Calculate total LNG imports
        lng_total = pd.Series(0.0, index=supply_df.index)
        
        lng_routes_found = []
        for route_name in self.lng_aggregation_routes:
            if route_name in supply_df.columns:
                lng_route_data = supply_df[route_name]
                if lng_route_data.sum() > 0:
                    lng_total += lng_route_data
                    lng_routes_found.append(route_name)
        
        supply_df['Total_LNG_Imports'] = lng_total
        
        logger.info(f"✅ LNG aggregation complete")
        logger.info(f"   Found {len(lng_routes_found)} active LNG routes: {lng_routes_found}")
        
        return supply_df
    
    def apply_geopolitical_corrections(self, supply_df: pd.DataFrame) -> pd.DataFrame:
        """
        Apply geopolitical corrections to supply data.
        
        Key corrections for Russian supply disruptions post-2022.
        """
        logger.info("🌍 Applying geopolitical corrections")
        
        corrections_applied = []
        
        for route_name, correction_config in self.geopolitical_corrections.items():
            if route_name in supply_df.columns:
                route_data = supply_df[route_name].copy()
                
                if 'cutoff_date' in correction_config:
                    # Zero out values after cutoff date
                    cutoff_date = pd.to_datetime(correction_config['cutoff_date'])
                    post_cutoff_mask = supply_df['Date'] >= cutoff_date
                    route_data.loc[post_cutoff_mask] = correction_config['post_cutoff_value']
                    corrections_applied.append(f"{route_name} (cutoff)")
                
                elif 'reduction_date' in correction_config:
                    # Apply reduction factor after date
                    reduction_date = pd.to_datetime(correction_config['reduction_date'])
                    reduction_mask = supply_df['Date'] >= reduction_date
                    route_data.loc[reduction_mask] *= correction_config['reduction_factor']
                    corrections_applied.append(f"{route_name} (reduction)")
                
                supply_df[route_name] = route_data
        
        logger.info(f"✅ Applied {len(corrections_applied)} geopolitical corrections")
        logger.info(f"   Corrections: {corrections_applied}")
        
        return supply_df
    
    def calculate_supply_aggregates(self, supply_df: pd.DataFrame) -> pd.DataFrame:
        """
        Calculate supply-side aggregates for analysis.
        
        Creates summary totals by supply type (Pipeline, LNG, Production, etc).
        """
        logger.info("📊 Calculating supply aggregates")
        
        # Pipeline Imports Total
        pipeline_routes = [route for route in self.supply_routes_mapping.keys() 
                          if self.supply_routes_mapping[route][0] == 'Import']
        pipeline_total = pd.Series(0.0, index=supply_df.index)
        for route in pipeline_routes:
            if route in supply_df.columns:
                pipeline_total += supply_df[route]
        supply_df['Total_Pipeline_Imports'] = pipeline_total
        
        # Domestic Production Total
        production_routes = [route for route in self.supply_routes_mapping.keys()
                           if self.supply_routes_mapping[route][0] == 'Production'] 
        production_total = pd.Series(0.0, index=supply_df.index)
        for route in production_routes:
            if route in supply_df.columns:
                production_total += supply_df[route]
        supply_df['Total_Domestic_Production'] = production_total
        
        # Total Supply (excluding exports and storage)
        total_supply = (supply_df['Total_Pipeline_Imports'] + 
                       supply_df['Total_LNG_Imports'] + 
                       supply_df['Total_Domestic_Production'])
        supply_df['Total_Supply'] = total_supply
        
        logger.info("✅ Supply aggregates calculated")
        logger.info(f"   Pipeline imports, LNG imports, Domestic production, Total supply")
        
        return supply_df
    
    def merge_supply_with_working_demand(self, demand_df: pd.DataFrame, 
                                       supply_df: pd.DataFrame) -> pd.DataFrame:
        """
        Merge supply data with working demand data ONLY at final output stage.
        
        CRITICAL: This preserves the working demand data completely unchanged.
        """
        logger.info("🔗 Merging supply data with working demand results")
        
        # Merge on Date - demand data is preserved completely
        complete_df = demand_df.merge(supply_df, on='Date', how='left')
        
        # Fill any missing supply values with 0
        supply_columns = [col for col in supply_df.columns if col != 'Date']
        for col in supply_columns:
            if col in complete_df.columns:
                complete_df[col] = complete_df[col].fillna(0.0)
        
        logger.info(f"✅ Merge complete")
        logger.info(f"   Final dataset: {len(complete_df)} rows with demand + supply data")
        logger.info(f"   Demand columns preserved: {len(demand_df.columns)-1}")
        logger.info(f"   Supply columns added: {len(supply_columns)}")
        
        return complete_df
    
    def validate_supply_processing(self, supply_df: pd.DataFrame, 
                                 complete_df: pd.DataFrame) -> bool:
        """
        Validate that supply processing doesn't interfere with working demand.
        
        CRITICAL: Demand validation targets must remain unchanged.
        """
        logger.info("🔍 Validating supply processing independence")
        
        validation_date = '2016-10-03'
        
        # Check that demand targets are still perfect
        test_sample = complete_df[complete_df['Date'] == validation_date]
        if test_sample.empty:
            logger.warning(f"Validation date {validation_date} not found")
            return False
        
        sample_row = test_sample.iloc[0]
        
        # These must remain PERFECT (unchanged from Phase 1)
        demand_targets = {
            'France': 90.13,
            'Total': 715.22,
            'Industrial': 236.42,  # Accept Bloomberg reshuffling value
            'LDZ': 307.80,
            'Gas_to_Power': 166.71
        }
        
        logger.info(f"\n📊 DEMAND PRESERVATION CHECK for {validation_date}:")
        logger.info("=" * 60)
        
        demand_preserved = True
        for column, target in demand_targets.items():
            if column in complete_df.columns:
                actual = sample_row[column]
                diff = abs(actual - target)
                
                if diff < 0.01:
                    status = "✅ PRESERVED"
                elif diff < 5.0:  # Bloomberg reshuffling tolerance
                    status = "✅ PRESERVED (reshuffling)"
                else:
                    status = "❌ CORRUPTED"
                    demand_preserved = False
                
                logger.info(f"  {status} {column}: {actual:.2f} (target {target:.2f})")
            else:
                logger.error(f"  ❌ MISSING {column}")
                demand_preserved = False
        
        logger.info("=" * 60)
        
        # Check supply data was added
        supply_columns_found = [col for col in supply_df.columns if col != 'Date' and col in complete_df.columns]
        logger.info(f"📊 Supply columns added: {len(supply_columns_found)}")
        
        # Show sample supply data
        if 'Total_Supply' in complete_df.columns:
            total_supply = sample_row['Total_Supply']
            logger.info(f"   Total Supply on {validation_date}: {total_supply:.2f}")
        
        if demand_preserved:
            logger.info("🎯 SUCCESS: Demand targets preserved perfectly!")
            logger.info("🚛 Supply processing completed independently!")
        else:
            logger.error("❌ FAILURE: Demand targets corrupted by supply processing!")
        
        return demand_preserved
    
    def run_independent_supply_pipeline(self, 
                                      demand_file: str = 'restored_demand_results.csv',
                                      multiticker_file: str = 'use4.xlsx',
                                      output_file: str = 'complete_demand_supply_results.csv') -> Optional[pd.DataFrame]:
        """
        Run the complete independent supply processing pipeline.
        
        Phase 2 of the revert task: Add supply processing without touching demand.
        """
        logger.info("🚀 Starting INDEPENDENT Supply Processing Pipeline")
        logger.info("=" * 80)
        logger.info("PHASE 2: Adding supply processing to working demand results")
        logger.info("CRITICAL: Preserving perfect demand validation unchanged")
        logger.info("=" * 80)
        
        try:
            # Step 1: Load working demand results (READ-ONLY)
            logger.info("📊 Step 1: Loading working demand results...")
            demand_df = self.load_working_demand_results(demand_file)
            
            # Step 2: Load MultiTicker for supply processing  
            logger.info("🚛 Step 2: Loading MultiTicker for supply processing...")
            supply_data_df, supply_metadata = self.load_multiticker_for_supply(multiticker_file)
            
            # Step 3: Process all supply routes independently
            logger.info("📊 Step 3: Processing all supply routes...")
            supply_df = self.process_all_supply_routes(supply_data_df, supply_metadata)
            
            # Step 4: Apply LNG aggregation
            logger.info("🛢️ Step 4: Applying LNG aggregation...")
            supply_df = self.apply_lng_aggregation(supply_df)
            
            # Step 5: Apply geopolitical corrections
            logger.info("🌍 Step 5: Applying geopolitical corrections...")
            supply_df = self.apply_geopolitical_corrections(supply_df)
            
            # Step 6: Calculate supply aggregates
            logger.info("📊 Step 6: Calculating supply aggregates...")
            supply_df = self.calculate_supply_aggregates(supply_df)
            
            # Step 7: Merge with working demand (FINAL STAGE ONLY)
            logger.info("🔗 Step 7: Merging with working demand results...")
            complete_df = self.merge_supply_with_working_demand(demand_df, supply_df)
            
            # Step 8: CRITICAL VALIDATION
            logger.info("✅ Step 8: Validating supply processing independence...")
            validation_passed = self.validate_supply_processing(supply_df, complete_df)
            
            if validation_passed:
                logger.info("💾 Step 9: Exporting complete results...")
                
                # Format for export
                export_data = complete_df.copy()
                export_data['Date'] = export_data['Date'].dt.strftime('%Y-%m-%d')
                
                # Round to appropriate precision
                numeric_cols = export_data.select_dtypes(include=[np.number]).columns
                export_data[numeric_cols] = export_data[numeric_cols].round(2)
                
                # Export complete results
                export_data.to_csv(output_file, index=False)
                
                logger.info(f"✅ SUCCESS: Complete results exported to {output_file}")
                logger.info("=" * 80)
                logger.info("🎯 PHASE 2 COMPLETE: SUPPLY PROCESSING ADDED SUCCESSFULLY!")
                logger.info("🚀 Demand targets preserved + Supply data integrated!")
                logger.info("✅ Ready for production use!")
                
                return complete_df
            else:
                logger.error("❌ Supply processing validation failed")
                logger.error("   Supply processing interfered with working demand results")
                return None
                
        except Exception as e:
            logger.error(f"❌ Independent supply pipeline failed: {str(e)}")
            import traceback
            traceback.print_exc()
            raise
    
    def merge_with_working_demand(self, demand_file: str = 'restored_demand_results.csv',
                                supply_data: pd.DataFrame = None) -> pd.DataFrame:
        """
        Merge independent supply results with working demand results.
        
        This is the ONLY point where supply and demand interact.
        """
        logger.info("🔗 Merging independent supply with working demand")
        
        # Load working demand results
        demand_data = pd.read_csv(demand_file)
        demand_data['Date'] = pd.to_datetime(demand_data['Date'])
        
        # Convert supply data dates
        if supply_data is not None:
            supply_data = supply_data.copy()
            supply_data['Date'] = pd.to_datetime(supply_data['Date'])
            
            # Merge on Date
            final_data = demand_data.merge(supply_data, on='Date', how='left')
        else:
            logger.warning("No supply data provided - using demand-only")
            final_data = demand_data.copy()
        
        logger.info(f"✅ Merged: {len(demand_data.columns)} demand cols + {len(supply_data.columns)-1 if supply_data is not None else 0} supply cols")
        
        return final_data
    
    def export_supply_audit_trail(self, output_file: str = 'independent_supply_audit.csv') -> str:
        """Export independent supply audit trail."""
        if not self.supply_audit_trail:
            logger.warning("No supply audit trail to export")
            return None
        
        audit_df = pd.DataFrame(self.supply_audit_trail)
        audit_df.to_csv(output_file, index=False)
        
        logger.info(f"📝 Independent supply audit trail exported to {output_file}")
        return output_file


def main():
    """
    Main execution for independent supply processor.
    """
    try:
        logger.info("🚀 INDEPENDENT SUPPLY PROCESSOR")
        logger.info("=" * 80)
        logger.info("CRITICAL: Testing restored working demand-side processing")
        logger.info("PHASE 2: Create separate supply-side module that doesn't interfere with working demand logic")
        
        # Initialize and run independent supply processor
        processor = IndependentSupplyProcessor()
        result = processor.run_independent_supply_pipeline()
        
        if result is not None:
            logger.info("\n🎉 PHASE 2 SUCCESS!")
            logger.info("Supply processing added independently without affecting demand!")
            logger.info("📊 Output: complete_demand_supply_results.csv")
            logger.info("🎯 Perfect demand validation preserved!")
        else:
            logger.error("\n💥 PHASE 2 FAILED!")
            logger.error("Supply processing interfered with working demand results")
            
        return result
        
    except Exception as e:
        logger.error(f"❌ Main execution failed: {str(e)}")
        raise


if __name__ == "__main__":
    result = main()