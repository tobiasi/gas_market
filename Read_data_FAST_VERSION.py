# -*- coding: utf-8 -*-
"""
Fast version of Read_data.py using optimized spreadsheet updates
Expected speedup: 5-10x faster for Excel operations
"""

import numpy as np
import os
import pandas as pd
import datetime 
import xlrd

# ... (same imports and functions as original)

# OPTIMIZED SPREADSHEET UPDATER CLASS
class SpreadsheetUpdater:
    """Keep workbook in memory and save once at the end"""
    
    def __init__(self, path: str):
        self.path = path
        try:
            import openpyxl as ox
            self.wb = ox.load_workbook(path)
        except FileNotFoundError:
            # Create new workbook if it doesn't exist
            self.wb = ox.Workbook()
        self.update_count = 0
        self.ox = ox
    
    def update(self, _df, startcol: int = 1, startrow: int = 1, 
               sheet_name: str = "TDSheet", update_index: bool = True, 
               update_columns: bool = False):
        """Fast update without saving"""
        # Create sheet if it doesn't exist
        if sheet_name not in self.wb.sheetnames:
            self.wb.create_sheet(sheet_name)
        
        ws = self.wb[sheet_name]
        
        # Convert to numpy arrays for fastest access
        data_array = _df.values
        
        # Batch write
        for row_idx in range(data_array.shape[0]):
            if update_index:
                ws.cell(row=startrow + row_idx, column=startcol - 1).value = _df.index[row_idx]
            
            for col_idx in range(data_array.shape[1]):
                ws.cell(row=startrow + row_idx, column=startcol + col_idx).value = data_array[row_idx, col_idx]
        
        if update_columns:
            for col in range(_df.shape[1]):
                ws.cell(row=1, column=startcol + col).value = _df.columns[col].strftime('%d-%m-%Y')
        
        self.update_count += 1
        print(f"✓ Updated sheet '{sheet_name}' ({_df.shape[0]}x{_df.shape[1]}) - {self.update_count} total updates")
    
    def save(self):
        """Save all updates at once"""
        print(f"💾 Saving {self.update_count} updates to {self.path}...")
        start_time = time.time()
        self.wb.save(self.path)
        elapsed = time.time() - start_time
        print(f"✅ Saved in {elapsed:.2f} seconds")
        self.update_count = 0
    
    def close(self):
        """Close and save"""
        self.save()
        self.wb.close()


#%% MODIFIED MAIN PROCESSING (same as original until here)
# ... all the same processing code ...

#%% FAST EXCEL WRITING SECTION
print("\n🚀 Starting FAST Excel updates...")
start_total = time.time()

# Create single updater instance 
updater = SpreadsheetUpdater(filename)

# Convert all dates to Excel format ONCE
full_data_out = full_data.copy()
full_data_out.index = to_excel_dates(pd.to_datetime(full_data_out.index))

ldz_out = ldz.copy() 
gtp_out = gtp.copy()
industry_out = industry.copy()
countries_out = countries.copy()
supply_out = supply.copy()
lng_out = lng.copy()

ldz_out.index = to_excel_dates(pd.to_datetime(ldz_out.index))
gtp_out.index = to_excel_dates(pd.to_datetime(gtp_out.index))
industry_out.index = to_excel_dates(pd.to_datetime(industry_out.index))
countries_out.index = to_excel_dates(pd.to_datetime(countries.index))
supply_out.index = to_excel_dates(pd.to_datetime(supply_out.index))
lng_out.index = to_excel_dates(pd.to_datetime(lng_out.index))

# BATCH ALL UPDATES (instead of saving after each one)
print("📊 Updating main data sheets...")
updater.update(full_data_out, 2, 8, 'Multiticker')
updater.update(ldz_out, 2, 4, 'LDZ demand') 
updater.update(gtp_out, 2, 4, 'Gas-to-Power demand') 
updater.update(industry_out, 2, 4, 'Industrial demand')
updater.update(countries_out, 2, 4, 'Demand') 
updater.update(supply_out, 2, 4, 'Supply') 
updater.update(lng_out, 2, 4, 'LNG imports by country')

print("📈 Updating analysis sheets...")
# Convert other dataframes with Excel dates
DD.index = to_excel_dates(pd.to_datetime(DD.index))
DD_perc.index = to_excel_dates(pd.to_datetime(DD_perc.index))
Actuals.index = to_excel_dates(pd.date_range('2022-01-01', freq='D', periods=365))
# ... etc for other dataframes

# Update all analysis sheets
updater.update(DD_m, 2, 3, 'Calendar years - monthly', update_index=False) 
updater.update(DD_perc_m, 2, 3, 'Calendar years, % - monthly', update_index=False)
updater.update(DD_long_m, 2, 20, 'Calendar years - monthly', update_index=True)
updater.update(DD_cum_m, 7, 20, 'Calendar years - monthly', update_index=False)
updater.update(DD_perc_long_m, 2, 20, 'Calendar years, % - monthly', update_index=True)
updater.update(Actuals, 2, 3, 'Calendar years actuals', update_index=False)
updater.update(DD, 2, 3, 'Calendar years', update_index=False)
updater.update(DD_perc, 2, 3, 'Calendar years, %', update_index=False)
updater.update(Actuals_t, 2, 3, 'Calendar years YOY %', update_index=False)

# SAVE EVERYTHING AT ONCE (this is where the magic happens!)
updater.save()

total_elapsed = time.time() - start_total
print(f"🎉 Total Excel operations completed in {total_elapsed:.2f} seconds")
print(f"📊 Updated {updater.update_count} sheets in a single save operation")

# Close the updater
updater.close()

# Continue with rest of original code (Balance.xlsx creation, etc.)
# ...