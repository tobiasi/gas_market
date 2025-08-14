import os
from openpyxl import load_workbook

def analyze_excel_structure(filename):
    """Analyze the structure of an Excel file"""
    print(f"\n{'='*50}")
    print(f"ANALYZING: {filename}")
    print(f"{'='*50}")
    
    try:
        # Try openpyxl first (for .xlsx files)
        if filename.endswith('.xlsx'):
            try:
                workbook = load_workbook(filename, read_only=True)
                sheet_names = workbook.sheetnames
                print(f"Number of sheets: {len(sheet_names)}")
                print("Sheet names:")
                for i, sheet in enumerate(sheet_names):
                    print(f"  {i+1}. {sheet}")
                
                # Try to get basic info about each sheet
                for sheet_name in sheet_names[:3]:  # Limit to first 3 sheets for analysis
                    print(f"\n--- SHEET: {sheet_name} ---")
                    try:
                        sheet = workbook[sheet_name]
                        max_row = sheet.max_row
                        max_col = sheet.max_column
                        print(f"Max dimensions: {max_row} rows x {max_col} columns")
                        
                        # Try to read some sample data
                        sample_data = []
                        for row in range(1, min(6, max_row + 1)):
                            row_data = []
                            for col in range(1, min(11, max_col + 1)):
                                cell_value = sheet.cell(row=row, column=col).value
                                if cell_value is not None:
                                    row_data.append(str(cell_value)[:20])  # Truncate long values
                                else:
                                    row_data.append("")
                            sample_data.append(row_data)
                        
                        print("Sample data (first 5 rows, first 10 columns):")
                        for i, row in enumerate(sample_data):
                            print(f"Row {i+1}: {row}")
                            
                    except Exception as e:
                        print(f"Error analyzing sheet {sheet_name}: {e}")
                
                workbook.close()
                return
                        
            except Exception as e:
                print(f"Openpyxl error: {e}")
        
        # If openpyxl fails, we'll work with what we have
        print("Could not analyze with openpyxl")
                
    except Exception as e:
        print(f"Error reading file {filename}: {e}")

# Analyze the original file
analyze_excel_structure("DNB Markets EUROPEAN GAS BALANCE.xlsx")

# Analyze the temp files
temp_files = [f for f in os.listdir('.') if f.endswith('_temp.xlsx')]
for temp_file in sorted(temp_files):
    analyze_excel_structure(temp_file)